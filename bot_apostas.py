import os
import io
import re
import logging
import threading
import psycopg2
import psycopg2.extras
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InputFile
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters
)

TOKEN         = "8790751046:AAG-AsvU3V-K5j4U8IOUQrpT6NXX8K3FcjU"
DATABASE_URL  = os.environ.get("DATABASE_URL", "postgresql://apostas_db_br3e_user:9Q8kF2084mtEmOESc09jc22ZR7nS5FLz@dpg-d6ub6lfafjfc7380et2g-a/apostas_db_br3e")
BANCA_INICIAL = 5000

CASAS = ["Bet365","Betano","SportingBet/Betboo","Novibet","Vaidebet/Betpix365","Betfast","BETesporte",
         "Betnacional","BetFair","Stake","Lottu","Esportes Da Sorte","Esportivabet",
         "Outra"]

# Mapeamento para normalizar casas antigas/variantes para o nome atual
MAPA_CASAS_GEMEAS = {
    "sportingbet":    "SportingBet/Betboo",
    "Sportingbet":    "SportingBet/Betboo",
    "SportingBet":    "SportingBet/Betboo",
    "betboo":         "SportingBet/Betboo",
    "Betboo":         "SportingBet/Betboo",
    "vaidebet":       "Vaidebet/Betpix365",
    "Vaidebet":       "Vaidebet/Betpix365",
    "betpix365":      "Vaidebet/Betpix365",
    "Betpix365":      "Vaidebet/Betpix365",
    # Donaldbet/Bet.bet
    "donaldbet":      "Donaldbet/Bet.bet",
    "Donaldbet":      "Donaldbet/Bet.bet",
    "DonaldBet":      "Donaldbet/Bet.bet",
    "bet.bet":        "Donaldbet/Bet.bet",
    "Bet.bet":        "Donaldbet/Bet.bet",
    "BET.BET":        "Donaldbet/Bet.bet",
}

ESPORTES = ["⚽ Futebol","🏀 Basquete","🎾 Tênis","🏒 Hóquei",
            "🏈 Futebol Americano","⚾ Beisebol","🥊 MMA/Boxe","🏐 Vôlei",
            "🏎️ F1","🎮 Esports","Outro"]

logging.basicConfig(level=logging.WARNING)

# Estados dos ConversationHandlers
DATA, HORARIO, DESCRICAO, ODD, STAKE, ESPORTE, CASA = range(7)
ATUALIZAR_ID, ATUALIZAR_RES                         = range(7, 9)
EDITAR_ID, EDITAR_CAMPO, EDITAR_VALOR, EDITAR_CASA  = range(9, 13)
MUDAR_UNIDADE_VALOR                                 = 13
RESULTADOS_MENU                                     = 14
GERAR_MENU                                          = 15

CANCELAR_BTN = "❌ Cancelar"

# ── TECLADOS ──────────────────────────────────────────────────────────────────
def teclado_menu():
    return ReplyKeyboardMarkup([
        ["📝 Nova aposta",   "⏳ Ver pendentes"],
        ["📈 Resultados",    "✏️ Editar aposta"],
        ["📊 Gerar Resultados", "⚙️ Mudar Unidade"],
    ], resize_keyboard=True)

def teclado_cancelar():
    return ReplyKeyboardMarkup([[CANCELAR_BTN]], resize_keyboard=True)

def teclado_resultados():
    return ReplyKeyboardMarkup([
        ["🏦 Por Casa", "🏅 Por Esporte"],
        ["📅 Por Mês",  "🔙 Voltar"],
    ], resize_keyboard=True)

def teclado_gerar():
    return ReplyKeyboardMarkup([
        ["📊 Gerar Dashboard", "📂 Gerar Dados"],
        ["🔙 Voltar"],
    ], resize_keyboard=True)

# ── BANCO DE DADOS ────────────────────────────────────────────────────────────
def conectar():
    return psycopg2.connect(DATABASE_URL)

def inicializar_db():
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS apostas (
                    id        SERIAL PRIMARY KEY,
                    data      DATE NOT NULL,
                    horario   VARCHAR(5),
                    descricao TEXT NOT NULL,
                    odd       NUMERIC(8,3) NOT NULL,
                    stake     NUMERIC(10,2) NOT NULL,
                    resultado VARCHAR(10) DEFAULT 'pendente',
                    casa      VARCHAR(50),
                    esporte   VARCHAR(50),
                    freebet   NUMERIC(10,2) DEFAULT 0,
                    unidade   NUMERIC(10,2) DEFAULT 50
                )
            """)
            cur.execute("ALTER TABLE apostas ADD COLUMN IF NOT EXISTS freebet NUMERIC(10,2) DEFAULT 0")
            cur.execute("ALTER TABLE apostas ADD COLUMN IF NOT EXISTS unidade NUMERIC(10,2) DEFAULT 50")
            cur.execute("ALTER TABLE apostas ADD COLUMN IF NOT EXISTS cashout_valor NUMERIC(10,2) DEFAULT 0")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS configuracoes (
                    chave  VARCHAR(50) PRIMARY KEY,
                    valor  TEXT NOT NULL
                )
            """)
            cur.execute("""
                INSERT INTO configuracoes (chave, valor) VALUES ('unidade_atual', '50')
                ON CONFLICT (chave) DO NOTHING
            """)
        conn.commit()
    # Normaliza casas com nomes antigos/errados automaticamente a cada deploy
    migrar_casas_gemeas()

def carregar():
    with conectar() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM apostas ORDER BY data ASC, horario ASC, id ASC")
            return [dict(r) for r in cur.fetchall()]

def inserir(a):
    unidade = get_unidade_atual()
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO apostas (data,horario,descricao,odd,stake,resultado,casa,esporte,freebet,unidade)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (a["data"],a["horario"],a["descricao"],a["odd"],a["stake"],
                  a["resultado"],a["casa"],a["esporte"],float(a.get("freebet") or 0),unidade))
            new_id = cur.fetchone()[0]
        conn.commit()
    return new_id

def atualizar_campo(id_aposta, campo, valor):
    if campo not in {"data","horario","descricao","odd","stake","resultado","casa","esporte","freebet","cashout_valor"}:
        return
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE apostas SET {campo}=%s WHERE id=%s", (valor, id_aposta))
        conn.commit()

def buscar_por_id(id_aposta):
    with conectar() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM apostas WHERE id=%s", (id_aposta,))
            row = cur.fetchone()
    return dict(row) if row else None

def deletar_aposta(id_aposta):
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM apostas WHERE id=%s", (id_aposta,))
        conn.commit()

def get_unidade_atual():
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT valor FROM configuracoes WHERE chave='unidade_atual'")
            row = cur.fetchone()
    return float(row[0]) if row else 50.0

def migrar_casas_gemeas():
    """Atualiza no banco todas as apostas com casas antigas para os nomes com gêmeas."""
    with conectar() as conn:
        with conn.cursor() as cur:
            atualizadas = 0
            for nome_antigo, nome_novo in MAPA_CASAS_GEMEAS.items():
                cur.execute("UPDATE apostas SET casa=%s WHERE LOWER(casa)=LOWER(%s)", (nome_novo, nome_antigo))
                atualizadas += cur.rowcount
        conn.commit()
    return atualizadas

async def cmd_migrar_casas(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    n = migrar_casas_gemeas()
    await update.message.reply_text(
        f"✅ *Migração concluída!*\n{n} aposta(s) atualizadas com os nomes novos das casas gêmeas.",
        parse_mode="Markdown", reply_markup=teclado_menu()
    )


def set_unidade_atual(valor):
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO configuracoes (chave, valor) VALUES ('unidade_atual', %s)
                ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor
            """, (str(valor),))
        conn.commit()

# ── HELPERS ───────────────────────────────────────────────────────────────────
def normalizar_casa(s):
    s = (s or "").strip()
    if not s: return "Sem casa"
    return MAPA_CASAS_GEMEAS.get(s, s.title() if s not in CASAS else s)

MAPA_ESPORTE = {
    # Futebol
    "futebol":              "⚽ Futebol",
    "soccer":               "⚽ Futebol",
    "football":             "⚽ Futebol",
    "⚽ futebol":           "⚽ Futebol",
    # Basquete
    "basquete":             "🏀 Basquete",
    "basketball":           "🏀 Basquete",
    "🏀 basquete":          "🏀 Basquete",
    # Tênis
    "tenis":                "🎾 Tênis",
    "tênis":                "🎾 Tênis",
    "tennis":               "🎾 Tênis",
    "🎾 tênis":             "🎾 Tênis",
    # Hóquei
    "hoquei":               "🏒 Hóquei",
    "hóquei":               "🏒 Hóquei",
    "ice hockey":           "🏒 Hóquei",
    "hockey":               "🏒 Hóquei",
    "🏒 hóquei":            "🏒 Hóquei",
    # Futebol Americano
    "futebol americano":    "🏈 Futebol Americano",
    "american football":    "🏈 Futebol Americano",
    "nfl":                  "🏈 Futebol Americano",
    "🏈 futebol americano": "🏈 Futebol Americano",
    # Beisebol
    "beisebol":             "⚾ Beisebol",
    "baseball":             "⚾ Beisebol",
    "⚾ beisebol":          "⚾ Beisebol",
    # MMA/Boxe
    "mma/boxe":             "🥊 MMA/Boxe",
    "mma":                  "🥊 MMA/Boxe",
    "boxe":                 "🥊 MMA/Boxe",
    "boxing":               "🥊 MMA/Boxe",
    "🥊 mma/boxe":          "🥊 MMA/Boxe",
    # Vôlei
    "volei":                "🏐 Vôlei",
    "vôlei":                "🏐 Vôlei",
    "volleyball":           "🏐 Vôlei",
    "🏐 vôlei":             "🏐 Vôlei",
    # F1
    "f1":                   "🏎️ F1",
    "formula 1":            "🏎️ F1",
    "fórmula 1":            "🏎️ F1",
    "formula1":             "🏎️ F1",
    "🏎️ f1":               "🏎️ F1",
    # Esports
    "esports":              "🎮 Esports",
    "esport":               "🎮 Esports",
    "e-sports":             "🎮 Esports",
    "e-sport":              "🎮 Esports",
    "🎮 esports":           "🎮 Esports",
    # Cricket
    "cricket":              "🏏 Cricket",
    "🏏 cricket":           "🏏 Cricket",
}

def normalizar_esporte(s):
    if not s: return ""
    return MAPA_ESPORTE.get(s.strip().lower(), s.strip())

def lucro_aposta(a):
    stake   = float(a["stake"])
    odd     = float(a["odd"])
    try:
        freebet = float(a.get("freebet") or 0)
    except:
        freebet = stake if str(a.get("freebet","")).strip().lower() == "sim" else 0.0
    freebet = min(freebet, stake)
    if a["resultado"] == "ganhou":
        return stake * (odd - 1)
    if a["resultado"] == "perdeu":
        return -(stake - freebet)
    if a["resultado"] == "void":
        cashout_valor = float(a.get("cashout_valor") or 0)
        if cashout_valor > 0:
            return cashout_valor - stake  # lucro real do cashout
    return 0.0

def eh_cashout(a):
    """Retorna True se a aposta void foi um cashout (tem cashout_valor > 0)."""
    return a["resultado"] == "void" and float(a.get("cashout_valor") or 0) > 0

async def voltar_menu(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("O que mais?", reply_markup=teclado_menu())
    return ConversationHandler.END

async def cancelar(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("❌ Cancelado.", reply_markup=teclado_menu())
    return ConversationHandler.END

# ── /start ────────────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text(
        "👋 *Gestor de Apostas*\n\nEscolha uma opção:",
        reply_markup=teclado_menu(), parse_mode="Markdown"
    )

# ── MENU PRINCIPAL ────────────────────────────────────────────────────────────
async def menu_botao(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text

    if txt == "⏳ Ver pendentes":    return await ver_pendentes(update, ctx)
    if txt == "⚙️ Mudar Unidade":    return await mudar_unidade_inicio(update, ctx)

# ── NOVA APOSTA ───────────────────────────────────────────────────────────────
async def nova_aposta_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    hoje = datetime.now().strftime("%d/%m/%Y")
    await update.message.reply_text(
        f"📅 *Data do jogo*\nHoje é *{hoje}* — mande *0* para confirmar ou digite outra:",
        reply_markup=teclado_cancelar(), parse_mode="Markdown"
    )
    return DATA

async def receber_data(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    if raw == "0":
        ctx.user_data["data"] = datetime.now().strftime("%Y-%m-%d")
    else:
        data = normalizar_data(raw)
        if not data:
            await update.message.reply_text("❌ Data inválida:")
            return DATA
        ctx.user_data["data"] = data
    agora = datetime.now().strftime("%H:%M")
    await update.message.reply_text(
        f"⏰ *Horário do jogo*\nAgora são *{agora}* — mande *0* para usar ou digite outro (HH:MM):",
        reply_markup=teclado_cancelar(), parse_mode="Markdown"
    )
    return HORARIO

def normalizar_data(raw: str):
    """Converte entrada flexível de data para YYYY-MM-DD.
    Aceita: 20 → dia 20 do mês atual | 20/05 → 20/05 do ano atual | 20/05/2026 ou 20/05/26
    """
    agora = datetime.now()
    raw = raw.strip()
    # Só dia: 20 → dia 20 do mês e ano atuais
    try:
        dia = int(raw)
        if 1 <= dia <= 31:
            try:
                return datetime(agora.year, agora.month, dia).strftime("%Y-%m-%d")
            except ValueError:
                return None
    except ValueError:
        pass
    # DD/MM → ano atual
    try:
        d = datetime.strptime(raw, "%d/%m")
        return datetime(agora.year, d.month, d.day).strftime("%Y-%m-%d")
    except ValueError:
        pass
    # DD/MM/AAAA ou DD/MM/AA
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

async def receber_horario(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    if raw == "0":
        ctx.user_data["horario"] = datetime.now().strftime("%H:%M")
    else:
        horario = normalizar_horario(raw)
        if not horario:
            await update.message.reply_text("❌ Horário inválido:")
            return HORARIO
        ctx.user_data["horario"] = horario
    await update.message.reply_text("🏷 *Descrição:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return DESCRICAO

def normalizar_horario(raw: str):
    """Converte entrada flexível de horário para HH:MM.
    Aceita: 15 → 15:00 | 15:3 → 15:30 | 153 → 15:30 | 1530 → 15:30 | 15:30 → 15:30
    """
    import re
    raw = raw.strip().replace(".", ":").replace(",", ":")
    # Formato HH:MM ou H:MM
    m = re.fullmatch(r"(\d{1,2}):(\d{1,2})", raw)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        # 15:3 → 15:30
        if m.group(2) and len(m.group(2)) == 1:
            mi = int(m.group(2)) * 10
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return f"{h:02d}:{mi:02d}"
        return None
    # Só hora: 15 → 15:00
    m = re.fullmatch(r"(\d{1,2})", raw)
    if m:
        h = int(m.group(1))
        if 0 <= h <= 23:
            return f"{h:02d}:00"
        return None
    # Formato compacto: 1530 → 15:30, 153 → 15:30, 900 → 09:00
    m = re.fullmatch(r"(\d{3,4})", raw)
    if m:
        s = m.group(1).zfill(4)
        h, mi = int(s[:2]), int(s[2:])
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return f"{h:02d}:{mi:02d}"
        return None
    return None

async def receber_descricao(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text.strip() == CANCELAR_BTN: return await cancelar(update, ctx)
    ctx.user_data["descricao"] = update.message.text.strip()
    await update.message.reply_text("🔢 *Odd:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return ODD

async def receber_odd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    try:
        ctx.user_data["odd"] = float(raw.replace(",", "."))
        await update.message.reply_text("💰 *Stake (R$):*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
        return STAKE
    except ValueError:
        await update.message.reply_text("❌ Odd inválida:")
        return ODD

async def receber_stake(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    try:
        ctx.user_data["stake"] = float(raw.replace(",", "."))
        teclado_esp = [[e] for e in ESPORTES] + [[CANCELAR_BTN]]
        await update.message.reply_text("🏅 *Esporte:*",
            reply_markup=ReplyKeyboardMarkup(teclado_esp, resize_keyboard=True, one_time_keyboard=True),
            parse_mode="Markdown")
        return ESPORTE
    except ValueError:
        await update.message.reply_text("❌ Stake inválido:")
        return STAKE

async def receber_esporte(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    if ctx.user_data.pop("aguardando_esporte_custom", False):
        ctx.user_data["esporte"] = raw
    elif raw == "Outro":
        await update.message.reply_text("Digite o nome do esporte:", reply_markup=teclado_cancelar())
        ctx.user_data["aguardando_esporte_custom"] = True
        return ESPORTE
    else:
        ctx.user_data["esporte"] = raw
    teclado_casa = [[c] for c in CASAS] + [[CANCELAR_BTN]]
    await update.message.reply_text("🏦 *Casa de aposta:*",
        reply_markup=ReplyKeyboardMarkup(teclado_casa, resize_keyboard=True, one_time_keyboard=True),
        parse_mode="Markdown")
    return CASA

async def receber_casa(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    casa = update.message.text.strip()
    if casa == CANCELAR_BTN: return await cancelar(update, ctx)
    if casa == "Outra":
        await update.message.reply_text("Digite o nome da casa:", reply_markup=teclado_cancelar())
        ctx.user_data["aguardando_casa_custom"] = True
        return CASA
    ctx.user_data.pop("aguardando_casa_custom", False)
    ctx.user_data["casa"] = casa.strip().title() if casa not in CASAS else casa

    new_id = inserir({
        "data": ctx.user_data["data"], "horario": ctx.user_data.get("horario",""),
        "descricao": ctx.user_data["descricao"], "odd": ctx.user_data["odd"],
        "stake": ctx.user_data["stake"], "resultado": "pendente",
        "casa": ctx.user_data["casa"], "esporte": ctx.user_data.get("esporte",""),
    })
    data_fmt = datetime.strptime(ctx.user_data["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
    hora_txt = f" às {ctx.user_data['horario']}" if ctx.user_data.get("horario") else ""
    await update.message.reply_text(
        f"✅ *Aposta #{new_id} salva!*\n\n"
        f"📅 {data_fmt}{hora_txt}\n🏅 {ctx.user_data.get('esporte','')}\n"
        f"🏷 {ctx.user_data['descricao']}\n🔢 Odd: {ctx.user_data['odd']}\n"
        f"💰 Stake: R$ {float(ctx.user_data['stake']):.2f}\n🏦 Casa: {ctx.user_data['casa']}",
        parse_mode="Markdown"
    )
    return await voltar_menu(update, ctx)

# ── VER PENDENTES ─────────────────────────────────────────────────────────────
async def ver_pendentes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas   = carregar()
    pendentes = sorted([a for a in apostas if a["resultado"] == "pendente"],
                       key=lambda a: (str(a["data"]), a.get("horario") or ""))
    if not pendentes:
        await update.message.reply_text("✅ Nenhuma aposta pendente!")
        return
    linhas = [f"⏳ *{len(pendentes)} apostas pendentes:*\n"]
    for a in pendentes:
        data_fmt = a["data"].strftime("%d/%m/%Y") if hasattr(a["data"],"strftime") else str(a["data"])[:10]
        hora = f" {a.get('horario','')}" if a.get("horario") else ""
        linhas.append(f"*#{a['id']}* {data_fmt}{hora} | odd {float(a['odd']):g} | R${float(a['stake']):.0f} | {a.get('casa','')}\n_{a['descricao']}_\n")
    await update.message.reply_text("\n".join(linhas), parse_mode="Markdown")

# ── RESULTADOS ────────────────────────────────────────────────────────────────
NOMES_MES_PT = {
    "janeiro":1,"jan":1,"fevereiro":2,"fev":2,"março":3,"mar":3,"abril":4,"abr":4,
    "maio":5,"mai":5,"junho":6,"jun":6,"julho":7,"jul":7,"agosto":8,"ago":8,
    "setembro":9,"set":9,"outubro":10,"out":10,"novembro":11,"nov":11,"dezembro":12,"dez":12,
}

async def resultados(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas = carregar()
    res     = [a for a in apostas if a["resultado"] in ("ganhou","perdeu") or eh_cashout(a)]
    if not res:
        await update.message.reply_text("Nenhuma aposta resolvida ainda.")
        return
    lucro_total = sum(lucro_aposta(a) for a in res)
    stake_total = sum(float(a["stake"]) for a in res)
    vitorias    = sum(1 for a in res if a["resultado"] == "ganhou")
    roi         = lucro_total / stake_total if stake_total else 0
    progressao  = lucro_total / BANCA_INICIAL
    sinal       = "+" if lucro_total >= 0 else ""
    emoji       = "📈" if lucro_total >= 0 else "📉"
    pendentes_ap = [a for a in apostas if a["resultado"] == "pendente"]
    stake_curso  = sum(float(a["stake"]) for a in pendentes_ap)
    unidade_atual = get_unidade_atual()
    # Lucro em unidades: usa a unidade registrada em cada aposta
    lucro_unidades = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in res)

    agora = datetime.now()
    hoje_date = agora.date()

    # Mês atual
    res_mes = [a for a in res if
               (a["data"] if isinstance(a["data"], type(hoje_date)) else
                datetime.strptime(str(a["data"])[:10], "%Y-%m-%d").date()
               ).replace(day=1) == hoje_date.replace(day=1)]
    lucro_mes   = sum(lucro_aposta(a) for a in res_mes)
    lucro_mes_u = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in res_mes)
    sinal_m = "+" if lucro_mes >= 0 else ""
    emoji_m = "🟢" if lucro_mes >= 0 else "🔴"

    # Semana atual (domingo a sábado)
    dias_desde_domingo = (agora.weekday() + 1) % 7
    inicio_semana_date = (agora - timedelta(days=dias_desde_domingo)).date()
    res_semana = [a for a in res if
                  (a["data"] if isinstance(a["data"], type(hoje_date)) else
                   datetime.strptime(str(a["data"])[:10], "%Y-%m-%d").date()
                  ) >= inicio_semana_date]
    lucro_semana   = sum(lucro_aposta(a) for a in res_semana)
    lucro_semana_u = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in res_semana)
    sinal_s = "+" if lucro_semana >= 0 else ""
    emoji_s = "🟢" if lucro_semana >= 0 else "🔴"

    NOMES_MES_LABEL = ["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    nome_mes_atual  = NOMES_MES_LABEL[agora.month]

    await update.message.reply_text(
        f"{emoji} *Resultados Gerais*\n\n"
        f"💰 Lucro Total: *{sinal}R$ {lucro_total:.2f}* ({lucro_unidades:+.2f}u)\n"
        f"📊 ROI: *{roi:+.1%}* | Progressão: *{progressao:+.2%}*\n"
        f"🏆 {vitorias}V / {len(res)-vitorias}D\n"
        f"⏳ Stake em curso: *R$ {stake_curso:.2f}* ({len(pendentes_ap)} ap)\n\n"
        f"{emoji_m} {nome_mes_atual}: *{sinal_m}R$ {lucro_mes:.2f}* ({lucro_mes_u:+.2f}u) | {len(res_mes)} ap\n"
        f"{emoji_s} Semana: *{sinal_s}R$ {lucro_semana:.2f}* ({lucro_semana_u:+.2f}u) | {len(res_semana)} ap",
        reply_markup=teclado_resultados(), parse_mode="Markdown"
    )
    return RESULTADOS_MENU

async def resultados_por_casa(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas = carregar()
    res     = [a for a in apostas if a["resultado"] in ("ganhou","perdeu") or eh_cashout(a)]
    casas   = {}
    for a in res:
        casa = normalizar_casa(a.get("casa"))
        if casa not in casas:
            casas[casa] = {"ap":0,"g":0,"stake":0.0,"lucro":0.0}
        c = casas[casa]; c["ap"] += 1; c["stake"] += float(a["stake"])
        c["lucro"] += lucro_aposta(a)
        if a["resultado"] == "ganhou": c["g"] += 1
    ordenadas = sorted(casas.items(), key=lambda x: x[1]["lucro"], reverse=True)
    linhas    = ["🏦 *Por Casa:*\n"]
    for nome, c in ordenadas:
        roi   = c["lucro"]/c["stake"] if c["stake"] else 0
        emoji = "🟢" if c["lucro"] >= 0 else "🔴"
        sinal = "+" if c["lucro"] >= 0 else ""
        linhas.append(f"{emoji} *{nome}*\n  {c['ap']} ap | {c['g']}V/{c['ap']-c['g']}D | {sinal}R$ {c['lucro']:.2f} | ROI: {roi:+.1%}\n")
    linhas.append("\nDigite uma data (DD/MM) ou 🔙 Voltar:")
    await update.message.reply_text("\n".join(linhas), reply_markup=teclado_resultados(), parse_mode="Markdown")

async def resultados_por_mes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas = carregar()
    res     = [a for a in apostas if a["resultado"] in ("ganhou","perdeu") or eh_cashout(a)]
    if not res:
        await update.message.reply_text("Nenhuma aposta resolvida ainda.", reply_markup=teclado_resultados())
        return

    from collections import defaultdict
    por_mes = defaultdict(list)
    for a in res:
        data_obj = a["data"] if hasattr(a["data"], "strftime") else datetime.strptime(str(a["data"])[:10], "%Y-%m-%d")
        chave = data_obj.strftime("%Y-%m")
        por_mes[chave].append(a)

    NOMES_MES = ["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    linhas = ["📅 *Resultados por Mês:*\n"]
    acum = 0
    for chave in sorted(por_mes.keys()):
        ap  = por_mes[chave]
        ano, mes = int(chave[:4]), int(chave[5:])
        lucro_m = sum(lucro_aposta(a) for a in ap)
        stake_m = sum(float(a["stake"]) for a in ap)
        g_m     = sum(1 for a in ap if a["resultado"] == "ganhou")
        roi_m   = lucro_m / stake_m if stake_m else 0
        acum   += lucro_m
        emoji_m = "🟢" if lucro_m >= 0 else "🔴"
        sinal_m = "+" if lucro_m >= 0 else ""
        # Lucro em unidades usando a unidade de cada aposta individualmente
        lucro_u = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in ap)
        linhas.append(
            f"{emoji_m} *{NOMES_MES[mes]}/{ano}*\n"
            f"  {len(ap)} ap | {g_m}V/{len(ap)-g_m}D | ROI {roi_m:+.1%}\n"
            f"  Lucro: *{sinal_m}R$ {lucro_m:.2f}* ({lucro_u:+.2f}u)\n"
        )
    sinal_ac = "+" if acum >= 0 else ""
    linhas.append(f"📊 *Acumulado: {sinal_ac}R$ {acum:.2f}*")
    linhas.append("\nDigite uma data (DD/MM), 🏦 *Por Casa* ou 🔙 *Voltar*:")
    await update.message.reply_text("\n".join(linhas), reply_markup=teclado_resultados(), parse_mode="Markdown")

async def resultados_por_mes_especifico(update: Update, ctx: ContextTypes.DEFAULT_TYPE, mes: int, ano: int = None):
    """Mostra resultado de um mês específico (ex: 'Janeiro', 'Abr/2025')."""
    apostas = carregar()
    res     = [a for a in apostas if a["resultado"] in ("ganhou","perdeu") or eh_cashout(a)]
    agora   = datetime.now()
    NOMES_MES = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                 "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    # Se não especificou ano: tenta ano atual; se vazio, tenta ano anterior
    anos_busca = [agora.year] if ano else [agora.year, agora.year - 1]
    if ano:
        anos_busca = [ano]
    for ano_b in anos_busca:
        do_mes = [a for a in res if
                  str(a["data"])[:7] == f"{ano_b}-{mes:02d}"]
        if do_mes:
            unidade_atual = get_unidade_atual()
            lucro_m   = sum(lucro_aposta(a) for a in do_mes)
            stake_m   = sum(float(a["stake"]) for a in do_mes)
            g_m       = sum(1 for a in do_mes if a["resultado"] == "ganhou")
            roi_m     = lucro_m / stake_m if stake_m else 0
            lucro_u_m = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in do_mes)
            emoji_m   = "🟢" if lucro_m >= 0 else "🔴"
            sinal_m   = "+" if lucro_m >= 0 else ""
            await update.message.reply_text(
                f"{emoji_m} *{NOMES_MES[mes]}/{ano_b}*\n\n"
                f"💰 Lucro: *{sinal_m}R$ {lucro_m:.2f}* ({lucro_u_m:+.2f}u)\n"
                f"📊 ROI: *{roi_m:+.1%}*\n"
                f"🏆 {g_m}V / {len(do_mes)-g_m}D | {len(do_mes)} apostas",
                reply_markup=teclado_resultados(), parse_mode="Markdown"
            )
            return
    label_ano = f"/{ano}" if ano else ""
    await update.message.reply_text(
        f"Nenhuma aposta resolvida em {NOMES_MES[mes]}{label_ano}.",
        reply_markup=teclado_resultados()
    )

async def resultados_por_esporte(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas = carregar()
    res     = [a for a in apostas if a["resultado"] in ("ganhou","perdeu") or eh_cashout(a)]
    if not res:
        await update.message.reply_text("Nenhuma aposta resolvida ainda.", reply_markup=teclado_resultados())
        return
    esportes = {}
    for a in res:
        esp = normalizar_esporte(a.get("esporte")) or "Sem esporte"
        if esp not in esportes:
            esportes[esp] = {"ap":0,"g":0,"stake":0.0,"lucro":0.0}
        e = esportes[esp]; e["ap"] += 1; e["stake"] += float(a["stake"])
        e["lucro"] += lucro_aposta(a)
        if a["resultado"] == "ganhou": e["g"] += 1
    ordenados = sorted(esportes.items(), key=lambda x: x[1]["lucro"], reverse=True)
    linhas = ["🏅 *Por Esporte:*\n"]
    for nome, e in ordenados:
        roi   = e["lucro"]/e["stake"] if e["stake"] else 0
        emoji = "🟢" if e["lucro"] >= 0 else "🔴"
        sinal = "+" if e["lucro"] >= 0 else ""
        linhas.append(f"{emoji} *{nome}*\n  {e['ap']} ap | {e['g']}V/{e['ap']-e['g']}D | {sinal}R$ {e['lucro']:.2f} | ROI: {roi:+.1%}\n")
    linhas.append("\nDigite uma data (DD/MM) ou 🔙 Voltar:")
    await update.message.reply_text("\n".join(linhas), reply_markup=teclado_resultados(), parse_mode="Markdown")

async def resultados_dia(update: Update, ctx: ContextTypes.DEFAULT_TYPE, raw: str) -> bool:
    agora = datetime.now()
    raw   = raw.strip()

    # Busca por ano: 4 dígitos ex: 2026
    if raw.isdigit() and len(raw) == 4:
        ano = int(raw)
        if 2000 <= ano <= 2099:
            apostas  = carregar()
            do_ano   = [a for a in apostas if str(a["data"])[:4] == str(ano) and (a["resultado"] in ("ganhou","perdeu") or eh_cashout(a))]
            if not do_ano:
                await update.message.reply_text(f"Nenhuma aposta resolvida em {ano}.", reply_markup=teclado_resultados())
                return True
            lucro_a  = sum(lucro_aposta(a) for a in do_ano)
            stake_a  = sum(float(a["stake"]) for a in do_ano)
            g = sum(1 for a in do_ano if a["resultado"] == "ganhou")
            roi = lucro_a / stake_a if stake_a else 0
            emoji = "🟢" if lucro_a >= 0 else "🔴"
            sinal = "+" if lucro_a >= 0 else ""
            await update.message.reply_text(
                f"{emoji} *{ano}* — {len(do_ano)} apostas | {g}V/{len(do_ano)-g}D\n"
                f"💰 {sinal}R$ {lucro_a:.2f} | ROI: {roi:+.1%}",
                reply_markup=teclado_resultados(), parse_mode="Markdown"
            )
            return True

    # Resolve data usando a mesma lógica do cadastro
    data_obj = None
    # 0 = hoje
    if raw == "0":
        data_obj = agora
    else:
        data_str_norm = normalizar_data(raw)
        if data_str_norm:
            data_obj = datetime.strptime(data_str_norm, "%Y-%m-%d")

    if not data_obj:
        return False

    data_str = data_obj.strftime("%Y-%m-%d")
    apostas  = carregar()
    do_dia   = [a for a in apostas if str(a["data"])[:10] == data_str
                and a["resultado"] in ("ganhou","perdeu","void")]
    if not do_dia:
        await update.message.reply_text(
            f"Nenhuma aposta resolvida em {data_obj.strftime('%d/%m/%Y')}.",
            reply_markup=teclado_resultados())
        return True
    lucro_dia   = sum(lucro_aposta(a) for a in do_dia)
    lucro_dia_u = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in do_dia)
    g  = sum(1 for a in do_dia if a["resultado"] == "ganhou")
    p  = sum(1 for a in do_dia if a["resultado"] == "perdeu")
    co = sum(1 for a in do_dia if eh_cashout(a))
    vo = sum(1 for a in do_dia if a["resultado"] == "void" and not eh_cashout(a))
    emoji = "🟢" if lucro_dia >= 0 else "🔴"
    sinal = "+" if lucro_dia >= 0 else ""
    resumo_str = f"{g}V {p}D"
    if co: resumo_str += f" {co}CO"
    if vo: resumo_str += f" {vo}Void"
    linhas = [f"{emoji} *{data_obj.strftime('%d/%m/%Y')}* — {resumo_str} — {sinal}R$ {lucro_dia:.2f} ({lucro_dia_u:+.2f}u)\n"]
    emojis_res = {"ganhou":"✅","perdeu":"❌"}
    for a in do_dia:
        l   = lucro_aposta(a)
        u_a = float(a.get("unidade") or 50)
        l_u = l / u_a
        s   = "+" if l >= 0 else ""
        if eh_cashout(a):
            cv = float(a.get("cashout_valor") or 0)
            linhas.append(f"💸 *#{a['id']}* odd {float(a['odd']):g} | R${float(a['stake']):.2f} → CO R${cv:.2f} ({s}R$ {l:.2f} / {l_u:+.2f}u)\n_{a['descricao']}_\n")
        elif a["resultado"] == "void":
            linhas.append(f"↩️ *#{a['id']}* odd {float(a['odd']):g} | R${float(a['stake']):.2f} → Void (R$ 0.00 / 0.00u)\n_{a['descricao']}_\n")
        else:
            linhas.append(f"{emojis_res.get(a['resultado'],'')} *#{a['id']}* odd {float(a['odd']):g} | R${float(a['stake']):.2f} → {s}R$ {l:.2f} ({l_u:+.2f}u)\n_{a['descricao']}_\n")
    linhas.append("Digite outra data, 🏦 *Por Casa* ou 🔙 *Voltar*:")
    await update.message.reply_text("\n".join(linhas), reply_markup=teclado_resultados(), parse_mode="Markdown")
    return True

async def resultados_resposta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handler das respostas dentro do ConversationHandler de resultados."""
    txt = update.message.text
    if txt == "🔙 Voltar":
        return await voltar_menu(update, ctx)
    if txt == "🏦 Por Casa":
        await resultados_por_casa(update, ctx)
        return RESULTADOS_MENU
    if txt == "📅 Por Mês":
        await resultados_por_mes(update, ctx)
        return RESULTADOS_MENU
    if txt == "🏅 Por Esporte":
        await resultados_por_esporte(update, ctx)
        return RESULTADOS_MENU
    # Filtro por nome de mês: "Janeiro", "jan", "Abril/2026", etc.
    txt_lower = txt.strip().lower()
    # Detecta "mes/ano" ou "mes ano" ou só "mes"
    ano_filtro = None
    nome_teste = txt_lower
    for sep in ("/", " ", "-"):
        if sep in txt_lower:
            partes = txt_lower.split(sep, 1)
            if partes[1].isdigit() and len(partes[1]) == 4:
                nome_teste = partes[0].strip()
                ano_filtro = int(partes[1])
                break
    mes_num = NOMES_MES_PT.get(nome_teste)
    if mes_num:
        await resultados_por_mes_especifico(update, ctx, mes_num, ano_filtro)
        return RESULTADOS_MENU
    if await resultados_dia(update, ctx, txt):
        return RESULTADOS_MENU
    await update.message.reply_text(
        "❓ Não reconheci. Use os botões, digite uma data (20/05), ano (2026) ou mês (Janeiro, Abr, Fev/2025).",
        reply_markup=teclado_resultados()
    )
    return RESULTADOS_MENU
CAMPOS_EDITAVEIS = ["data","horario","descricao","odd","stake","esporte","casa","resultado","cashout","freebet","deletar"]
CAMPOS_LABEL     = {
    "data":"📅 Data","horario":"⏰ Horário","descricao":"🏷 Descrição",
    "odd":"🔢 Odd","stake":"💰 Stake","esporte":"🏅 Esporte",
    "casa":"🏦 Casa","resultado":"📊 Resultado","cashout":"💸 Cashout","freebet":"🎁 Freebet (R$)","deletar":"🗑 Deletar aposta"
}

async def editar_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas   = carregar()
    pendentes = sorted([a for a in apostas if a["resultado"] == "pendente"],
                       key=lambda a: (str(a["data"]), a.get("horario") or ""))
    linhas = ["✏️ *Editar aposta*\n"]
    if pendentes:
        linhas.append("*Apostas pendentes:*\n")
        for a in pendentes:
            data_fmt = a["data"].strftime("%d/%m") if hasattr(a["data"],"strftime") else str(a["data"])[:10]
            hora = f" {a.get('horario','')}" if a.get("horario") else ""
            linhas.append(f"⏳ *#{a['id']}* {data_fmt}{hora} | odd {float(a['odd']):g} | _{str(a['descricao'])[:30]}_")
        linhas.append("")
    else:
        linhas.append("Nenhuma aposta pendente.\n")
    linhas.append("Digite o *ID* da aposta que quer editar:")
    await update.message.reply_text("\n".join(linhas), reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return EDITAR_ID

async def editar_receber_id(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    try:
        id_alvo = int(raw)
    except ValueError:
        await update.message.reply_text("❌ Digite só o número do ID:")
        return EDITAR_ID
    aposta = buscar_por_id(id_alvo)
    if not aposta:
        await update.message.reply_text("❌ ID não encontrado.")
        return EDITAR_ID
    ctx.user_data["editar_id"] = id_alvo
    data_fmt = aposta["data"].strftime("%d/%m/%Y") if hasattr(aposta["data"],"strftime") else str(aposta["data"])[:10]
    hora = f" {aposta.get('horario','')}" if aposta.get("horario") else ""
    _label_res = {"ganhou":"Green ✅","perdeu":"Red ❌","void":"Void ↩️","pendente":"Pendente ⏳"}
    info = (f"✏️ *Aposta #{id_alvo}*\n\n📅 {data_fmt}{hora}\n🏷 {aposta['descricao']}\n"
            f"🔢 Odd: {aposta['odd']}\n💰 Stake: R$ {float(aposta['stake']):.2f}\n"
            f"🏦 Casa: {aposta.get('casa','')}\n📊 Resultado: {_label_res.get(aposta['resultado'], aposta['resultado'])}\n🎁 Freebet: R$ {float(aposta['freebet'] or 0):.2f}\n\n*Qual campo editar?*")
    teclado = [[CAMPOS_LABEL[c]] for c in CAMPOS_EDITAVEIS] + [[CANCELAR_BTN]]
    await update.message.reply_text(info,
        reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True),
        parse_mode="Markdown")
    return EDITAR_CAMPO

async def editar_receber_campo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    campo = next((k for k,v in CAMPOS_LABEL.items() if v == raw), None)
    if not campo:
        await update.message.reply_text("❌ Escolha um dos campos:")
        return EDITAR_CAMPO
    ctx.user_data["editar_campo"] = campo
    if campo == "casa":
        teclado = [[c] for c in CASAS] + [[CANCELAR_BTN]]
        await update.message.reply_text("🏦 *Nova casa:*",
            reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True),
            parse_mode="Markdown")
        return EDITAR_CASA
    if campo == "esporte":
        teclado = [[e] for e in ESPORTES] + [[CANCELAR_BTN]]
        await update.message.reply_text("🏅 *Novo esporte:*",
            reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True),
            parse_mode="Markdown")
        return EDITAR_CASA
    if campo == "resultado":
        teclado = [["✅ Green","❌ Red"],["↩️ Void","⏳ Pendente"],[CANCELAR_BTN]]
        await update.message.reply_text("📊 *Qual o resultado?*",
            reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True),
            parse_mode="Markdown")
        return EDITAR_CASA
    if campo == "freebet":
        aposta_atual = buscar_por_id(ctx.user_data["editar_id"])
        stake_atual  = float(aposta_atual["stake"])
        await update.message.reply_text(
            f"🎁 *Freebet*\nDigite o valor em R$ que foi freebet nessa aposta.\n"
            f"Stake total: R$ {stake_atual:.2f}\n"
            f"Digite *0* para remover freebet:",
            reply_markup=teclado_cancelar(), parse_mode="Markdown")
        return EDITAR_VALOR
    if campo == "deletar":
        id_alvo  = ctx.user_data["editar_id"]
        aposta   = buscar_por_id(id_alvo)
        descr    = str(aposta["descricao"])[:30] if aposta else "?"
        await update.message.reply_text(
            f"🗑 *Deletar Aposta #{id_alvo}*\n_{descr}_\n\n"
            f"⚠️ Esta ação é *irreversível*. Digite *SIM* para confirmar:",
            reply_markup=teclado_cancelar(), parse_mode="Markdown")
        return EDITAR_VALOR
    if campo == "cashout":
        await update.message.reply_text(
            "💸 *Cashout*\nDigite o valor que voce *recebeu* de volta (R$):\n"
            "Ex: apostou R$50, fez cashout por R$30 -> digite *30*\n"
            "Se perdeu tudo -> digite *0*",
            reply_markup=teclado_cancelar(), parse_mode="Markdown")
        return EDITAR_VALOR
    dicas = {"data":"Nova data ou 0 para hoje:","horario":"Novo horário ou 0 para agora:",
             "descricao":"Nova descrição:","odd":"Nova odd:","stake":"Novo stake (R$):"}
    await update.message.reply_text(dicas[campo], reply_markup=teclado_cancelar())
    return EDITAR_VALOR

async def editar_receber_valor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw   = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    campo   = ctx.user_data["editar_campo"]
    id_alvo = ctx.user_data["editar_id"]
    novo_valor = None
    if campo == "deletar":
        if raw.upper() == "SIM":
            deletar_aposta(id_alvo)
            await update.message.reply_text(f"🗑 *Aposta #{id_alvo} deletada com sucesso!*", parse_mode="Markdown")
        else:
            await update.message.reply_text("❌ Cancelado. Aposta não deletada.")
        return await voltar_menu(update, ctx)
    if campo == "data":
        if raw == "0": novo_valor = datetime.now().strftime("%Y-%m-%d")
        else:
            novo_valor = normalizar_data(raw)
            if not novo_valor:
                await update.message.reply_text("❌ Data inválida:")
                return EDITAR_VALOR
    elif campo == "horario":
        if raw == "0":
            novo_valor = datetime.now().strftime("%H:%M")
        else:
            novo_valor = normalizar_horario(raw)
            if not novo_valor:
                await update.message.reply_text("❌ Horário inválido:")
                return EDITAR_VALOR
    elif campo in ("odd","stake","cashout","freebet"):
        try: novo_valor = float(raw.replace(",","."))
        except ValueError:
            await update.message.reply_text("❌ Digite um número:")
            return EDITAR_VALOR
        if campo == "freebet":
            aposta_atual = buscar_por_id(id_alvo)
            stake_atual  = float(aposta_atual["stake"])
            if novo_valor > stake_atual:
                await update.message.reply_text(
                    f"❌ Freebet não pode ser maior que o stake (R$ {stake_atual:.2f}). Digite novamente:")
                return EDITAR_VALOR
        if campo == "cashout":
            aposta   = buscar_por_id(id_alvo)
            stake    = float(aposta["stake"])
            recebido = novo_valor
            lucro_co = recebido - stake
            agora    = datetime.now()
            atualizar_campo(id_alvo, "resultado",     "void")
            atualizar_campo(id_alvo, "cashout_valor", recebido)
            atualizar_campo(id_alvo, "data",          agora.strftime("%Y-%m-%d"))
            atualizar_campo(id_alvo, "horario",       agora.strftime("%H:%M"))
            sinal    = "+" if lucro_co >= 0 else ""
            emoji_co = "📈" if lucro_co >= 0 else "📉"
            await update.message.reply_text(
                f"💸 *Cashout registrado!*\n\n"
                f"💰 Apostado: R$ {stake:.2f}\n"
                f"💸 Recebido: R$ {recebido:.2f}\n"
                f"{emoji_co} Resultado: *{sinal}R$ {lucro_co:.2f}*\n"
                f"📅 Planilhado em: *{agora.strftime('%d/%m/%Y às %H:%M')}*",
                parse_mode="Markdown")
            return await voltar_menu(update, ctx)
    else:
        novo_valor = raw
    return await aplicar_edicao(update, ctx, id_alvo, campo, novo_valor)

async def editar_receber_casa(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw   = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    campo   = ctx.user_data["editar_campo"]
    id_alvo = ctx.user_data["editar_id"]
    if campo == "resultado":
        mapa = {"✅ Green":"ganhou","❌ Red":"perdeu","✅ Ganhou":"ganhou","❌ Perdeu":"perdeu","↩️ Void":"void","⏳ Pendente":"pendente"}
        valor = mapa.get(raw, raw)
        return await aplicar_edicao(update, ctx, id_alvo, "resultado", valor)

    if raw == "Outra" or raw == "Outro":
        await update.message.reply_text("Digite o nome:", reply_markup=teclado_cancelar())
        ctx.user_data["editar_campo"] = campo + "_custom"
        return EDITAR_VALOR
    return await aplicar_edicao(update, ctx, id_alvo, campo, raw)

async def aplicar_edicao(update, ctx, id_alvo, campo, novo_valor):
    campo_real = campo.replace("_custom","")
    if campo_real == "casa" and isinstance(novo_valor, str) and novo_valor not in CASAS:
        novo_valor = novo_valor.strip().title()
    atualizar_campo(id_alvo, campo_real, novo_valor)
    label = CAMPOS_LABEL.get(campo_real, campo_real)
    exibe = novo_valor
    if campo_real == "data" and isinstance(novo_valor,str) and len(novo_valor)==10:
        exibe = datetime.strptime(novo_valor,"%Y-%m-%d").strftime("%d/%m/%Y")
    elif campo_real in ("odd","stake"):
        exibe = f"{float(novo_valor):.2f}"
    # Escolher emoji baseado no resultado se campo for resultado
    emoji_conf = "✅"
    if campo_real == "resultado":
        if novo_valor == "perdeu": emoji_conf = "❌"
        elif novo_valor == "void": emoji_conf = "↩️"
        elif novo_valor == "pendente": emoji_conf = "⏳"
        exibe = {"ganhou":"Green ✅","perdeu":"Red ❌","void":"Void ↩️","pendente":"Pendente ⏳"}.get(novo_valor, novo_valor)
    await update.message.reply_text(f"{emoji_conf} *Aposta #{id_alvo} atualizada!*\n{label} → *{exibe}*", parse_mode="Markdown")
    return await voltar_menu(update, ctx)


# ── MUDAR UNIDADE ─────────────────────────────────────────────────────────────
async def mudar_unidade_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    unidade_atual = get_unidade_atual()
    await update.message.reply_text(
        f"⚙️ *Mudar Unidade*\n\n"
        f"Unidade atual: *R$ {unidade_atual:.2f}*\n\n"
        f"A nova unidade será salva em todas as apostas *a partir de agora*.\n"
        f"Apostas anteriores mantêm a unidade original no histórico.\n\n"
        f"Digite o novo valor da unidade em R$:",
        reply_markup=teclado_cancelar(), parse_mode="Markdown"
    )
    return MUDAR_UNIDADE_VALOR

async def mudar_unidade_receber(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN: return await cancelar(update, ctx)
    try:
        novo_valor = float(raw.replace(",", "."))
        if novo_valor <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Digite um valor válido maior que zero:")
        return MUDAR_UNIDADE_VALOR
    unidade_anterior = get_unidade_atual()
    set_unidade_atual(novo_valor)
    await update.message.reply_text(
        f"✅ *Unidade atualizada!*\n\n"
        f"Anterior: R$ {unidade_anterior:.2f}\n"
        f"Nova: *R$ {novo_valor:.2f}*\n\n"
        f"Todas as apostas daqui pra frente usarão essa unidade.",
        parse_mode="Markdown"
    )
    return await voltar_menu(update, ctx)

# ── GERAR MENU ────────────────────────────────────────────────────────────────
async def gerar_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📊 *Exportar*\n\n"
        "• *Gerar Dashboard* → xlsx com estatísticas e gráficos\n"
        "• *Gerar Dados* → xlsx com dados brutos para migração de banco",
        reply_markup=teclado_gerar(), parse_mode="Markdown"
    )
    return GERAR_MENU

async def gerar_resposta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handler das respostas dentro do ConversationHandler de gerar."""
    txt = update.message.text
    if txt == "🔙 Voltar":
        return await voltar_menu(update, ctx)
    if txt == "📊 Gerar Dashboard":
        await gerar_dashboard(update, ctx)
        await update.message.reply_text("O que mais?", reply_markup=teclado_gerar())
        return GERAR_MENU
    if txt == "📂 Gerar Dados":
        await gerar_xlsx_dados(update, ctx)
        return GERAR_MENU
    await update.message.reply_text("❓ Use os botões abaixo.", reply_markup=teclado_gerar())
    return GERAR_MENU

async def gerar_xlsx_dados(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Gera o migrar_db.py pronto para uso no próximo mês."""
    await update.message.reply_text("⏳ Gerando script de migração, aguarde...")
    try:
        apostas = carregar()
        unidade_atual = get_unidade_atual()
        max_id = max((a["id"] for a in apostas), default=0)
        total = len(apostas)
        from datetime import datetime as dt2

        script = f'''"""
Migração de banco de dados — Bot Apostas
Gerado em {dt2.now().strftime("%d/%m/%Y %H:%M")} | {total} apostas | unidade atual: R$ {unidade_atual:.0f}

Como usar:
  1. Troque NEW_URL pela Internal URL do banco novo (Render → banco novo → Connections → Internal Database URL)
  2. Render → suspenda o banco atual → ative o banco novo
  3. No CMD:
       cd C:\\\\Users\\\\enzop\\\\Downloads
       python migrar_db.py
  4. Render → seu serviço → Environment → DATABASE_URL
     Troque pela Internal URL do banco novo → Save Changes → Manual Deploy
"""
import psycopg2
import psycopg2.extras

OLD_URL = "{DATABASE_URL}"
NEW_URL = "COLE_AQUI_A_INTERNAL_URL_DO_BANCO_NOVO"  # <-- troque isso

if __name__ == "__main__":
    if NEW_URL == "COLE_AQUI_A_INTERNAL_URL_DO_BANCO_NOVO":
        print("❌ Troque NEW_URL pela Internal URL do novo banco antes de rodar.")
        exit(1)

    print("Conectando ao banco atual...")
    conn_old = psycopg2.connect(OLD_URL)
    cur_old  = conn_old.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur_old.execute("SELECT * FROM apostas ORDER BY id ASC")
    apostas  = [dict(r) for r in cur_old.fetchall()]
    cur_old.execute("SELECT * FROM configuracoes")
    configs  = [dict(r) for r in cur_old.fetchall()]
    cur_old.execute("SELECT valor FROM configuracoes WHERE chave=\'unidade_atual\'")
    row_u    = cur_old.fetchone()
    unidade  = row_u["valor"] if row_u else "{unidade_atual:.0f}"
    conn_old.close()
    print(f"  {{len(apostas)}} apostas lidas | unidade atual: R$ {{unidade}}")

    print("Conectando ao banco novo...")
    conn_new = psycopg2.connect(NEW_URL)
    cur_new  = conn_new.cursor()
    cur_new.execute("""CREATE TABLE IF NOT EXISTS apostas (
        id SERIAL PRIMARY KEY, data DATE NOT NULL, horario VARCHAR(5),
        descricao TEXT NOT NULL, odd NUMERIC(8,3) NOT NULL, stake NUMERIC(10,2) NOT NULL,
        resultado VARCHAR(10) DEFAULT \'pendente\', casa VARCHAR(50), esporte VARCHAR(50),
        freebet NUMERIC(10,2) DEFAULT 0, unidade NUMERIC(10,2) DEFAULT 50,
        cashout_valor NUMERIC(10,2) DEFAULT 0)""")
    cur_new.execute("""CREATE TABLE IF NOT EXISTS configuracoes (
        chave VARCHAR(50) PRIMARY KEY, valor TEXT NOT NULL)""")
    cur_new.execute("DELETE FROM apostas")
    cur_new.execute("DELETE FROM configuracoes")

    for a in apostas:
        cur_new.execute("""INSERT INTO apostas
            (id,data,horario,descricao,odd,stake,resultado,casa,esporte,freebet,unidade,cashout_valor)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (a["id"],a["data"],a.get("horario"),a["descricao"],a["odd"],a["stake"],
             a["resultado"],a.get("casa"),a.get("esporte"),
             float(a.get("freebet") or 0), float(a.get("unidade") or 50),
             float(a.get("cashout_valor") or 0)))

    for c in configs:
        cur_new.execute("""INSERT INTO configuracoes (chave,valor) VALUES (%s,%s)
            ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor""",
            (c["chave"], c["valor"]))
    cur_new.execute("""INSERT INTO configuracoes (chave,valor) VALUES (\'unidade_atual\',%s)
        ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor""", (unidade,))

    if apostas:
        cur_new.execute("SELECT setval(\'apostas_id_seq\', %s)", (max(a["id"] for a in apostas),))

    conn_new.commit(); conn_new.close()
    print(f"\\n✅ {{len(apostas)}} apostas migradas | unidade: R$ {{unidade}}")
    print("\\nAgora faça o Manual Deploy no Render com a nova DATABASE_URL.")
'''

        script_bytes = script.encode("utf-8")
        instrucoes = (
            "📂 *Script de migração gerado!*\n\n"
            f"_{total} apostas | unidade atual: R$ {unidade_atual:.0f}_\n\n"
            "*Como usar no próximo mês:*\n\n"
            "1️⃣ Salve o `migrar\\_db.py` na pasta Downloads\n"
            "2️⃣ Abra o arquivo e troque `NEW\\_URL` pela *Internal URL* do banco novo\n"
            "3️⃣ Render → suspenda o banco atual → ative o banco novo\n"
            "4️⃣ No CMD:\n"
            "   `cd C:\\\\Users\\\\enzop\\\\Downloads`\n"
            "   `python migrar\\_db.py`\n"
            "5️⃣ Render → seu serviço → Environment → `DATABASE\\_URL`\n"
            "   Troque pela *Internal URL* do banco novo → *Manual Deploy*"
        )
        await update.message.reply_document(
            document=InputFile(io.BytesIO(script_bytes), filename="migrar_db.py"),
            caption=instrucoes,
            parse_mode="Markdown",
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Erro:\n`{str(e)}`", parse_mode="Markdown")

# ── GERAR DASHBOARD ───────────────────────────────────────────────────────────
async def gerar_dashboard(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⏳ Gerando dashboard, aguarde...")
    try:
        await _gerar_dashboard_interno(update, ctx)
    except Exception as _e:
        import traceback as _tb
        _trace = _tb.format_exc()[-700:]
        await update.message.reply_text(
            f"❌ Erro ao gerar dashboard:\n`{str(_e)}`\n\n`{_trace}`",
            parse_mode="Markdown", reply_markup=teclado_gerar()
        )

async def _gerar_dashboard_interno(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    from datetime import timedelta

    apostas = carregar()
    apostas_ord = sorted(apostas, key=lambda a: (str(a["data"]), a.get("horario") or "99:99", int(a["id"])))
    df_res = [a for a in apostas_ord if a["resultado"] in ("ganhou","perdeu") or eh_cashout(a)]

    BANCA = 5000
    DARK  = "1E293B"; GREEN = "16A34A"; RED = "DC2626"; AMBER = "D97706"
    WHITE = "FFFFFF"; ALT   = "EFF6FF"; BORDER_C = "CBD5E1"

    def est(cell, bold=False, fc="FFFFFF", bg=None, size=11, align="center"):
        cell.font = Font(name="Arial", bold=bold, color=fc, size=size)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if bg: cell.fill = PatternFill("solid", start_color=bg)

    def brd():
        s = Side(style="thin", color=BORDER_C)
        return Border(left=s, right=s, top=s, bottom=s)

    def cor(val):
        try: return GREEN if float(val) >= 0 else RED
        except: return "000000"

    # Métricas
    lucro_total = sum(
        lucro_aposta(a) for a in df_res
    )
    stake_total = sum(float(a["stake"]) for a in df_res)
    vitorias    = sum(1 for a in df_res if a["resultado"]=="ganhou")
    win_rate    = vitorias/len(df_res) if df_res else 0
    roi         = lucro_total/stake_total if stake_total else 0
    progressao  = lucro_total/BANCA
    total       = len(apostas)
    resolvidas  = len(df_res)
    pendentes   = total - resolvidas

    # Banca acumulada
    banca_acum = []
    acum = 0
    for a in df_res:
        acum += lucro_aposta(a)
        banca_acum.append(acum)

    wb = openpyxl.Workbook()

    # ── ABA 1: DASHBOARD ──
    ws = wb.active; ws.title = "Dashboard"
    ws.row_dimensions[2].height = 8

    unidade_dash = get_unidade_atual()
    lucro_units_dash = sum(lucro_aposta(a) / float(a.get("unidade") or 50) for a in df_res)
    stake_curso = sum(float(a["stake"]) for a in apostas if a["resultado"]=="pendente")
    cards = [
        ("Total",         str(total),                    DARK),
        ("Resolvidas",    str(resolvidas),               DARK),
        ("Pendentes",     str(pendentes),                AMBER if pendentes else DARK),
        ("Win Rate",      f"{win_rate:.1%}",             GREEN if win_rate>=0.5 else RED),
        ("Lucro R$",      f"R${lucro_total:.2f}",        GREEN if lucro_total>=0 else RED),
        ("Lucro Units",   f"{lucro_units_dash:+.1f}u",   GREEN if lucro_units_dash>=0 else RED),
        ("ROI",           f"{roi:+.1%}",                 GREEN if roi>=0 else RED),
        ("Progressao",    f"{progressao:+.1%}",          GREEN if progressao>=0 else RED),
        ("Em Curso",      f"R${stake_curso:.0f}",        AMBER if stake_curso>0 else DARK),
    ]
    # 9 cards em 9 colunas com largura uniforme
    for i,(label,val,bg) in enumerate(cards,1):
        c=ws.cell(row=3,column=i,value=label); est(c,bg=bg,size=9,fc="DBEAFE"); ws.row_dimensions[3].height=20
        v=ws.cell(row=4,column=i,value=val);   est(v,bold=True,bg=bg,size=12);  ws.row_dimensions[4].height=32
        d=ws.cell(row=5,column=i);              est(d,bg="0F172A");              ws.row_dimensions[5].height=4
        ws.column_dimensions[get_column_letter(i)].width=15
    ws.row_dimensions[6].height=10

    HDR=7; DAT=8
    headers=["#","Data","Hora","Descricao","Odd","Stake","Esporte","Casa","Resultado","Lucro","Banca","Progressao"]
    ws.row_dimensions[HDR].height=22
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=HDR,column=c,value=h); est(cell,bold=True,bg=DARK,size=10); cell.border=brd()

    acum2=0
    for i,a in enumerate(apostas_ord):
        er=DAT+i; ws.row_dimensions[er].height=18
        rb=WHITE if i%2==0 else ALT
        res=a["resultado"]
        lucro_a=banca_a=prog_a=""
        if res in ("ganhou","perdeu") or eh_cashout(a):
            lucro_a=lucro_aposta(a)
            acum2+=lucro_a; banca_a=round(acum2,2); prog_a=round(acum2/BANCA,4)
        res_d={"ganhou":"Green","perdeu":"Red","void":"Void","pendente":"Pendente"}.get(res,res)
        if eh_cashout(a): res_d="Cashout"
        data_fmt=a["data"].strftime("%d/%m/%Y") if hasattr(a["data"],"strftime") else str(a["data"])[:10]
        vals=[a["id"],data_fmt,a.get("horario",""),a["descricao"],a["odd"],a["stake"],
              a.get("esporte",""),a.get("casa",""),res_d,lucro_a,banca_a,prog_a]
        for c,val in enumerate(vals,1):
            cell=ws.cell(row=er,column=c,value=val)
            fc="000000"
            if c in (10,11) and val!="": fc=cor(val)
            if c==12 and val!="": fc=cor(val)
            cell.font=Font(name="Arial",size=10,color=fc)
            cell.alignment=Alignment(horizontal="left" if c==4 else "center",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if c in (5,6,10,11) and val!="": cell.number_format="#,##0.00"
            if c==12 and val!="": cell.number_format="+0.00%;-0.00%;0.00%"

    tr=DAT+len(apostas_ord); ws.row_dimensions[tr].height=20
    for c in range(1,13):
        cell=ws.cell(row=tr,column=c)
        if c==4: cell.value="TOTAL"
        elif c==6: cell.value=f"=SUM(F{DAT}:F{tr-1})"; cell.number_format="#,##0.00"
        elif c==10: cell.value=f"=SUM(J{DAT}:J{tr-1})"; cell.number_format="#,##0.00"
        est(cell,bold=True,bg=DARK,size=10); cell.border=brd()

    widths=[8,12,11,32,14,12,14,14,13,12,13,12]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.merge_cells("A1:L1"); ws["A1"] = "DASHBOARD DE APOSTAS"
    est(ws["A1"], bold=True, bg=DARK, size=16); ws.row_dimensions[1].height=40

    # ── ABA 2: ESTATÍSTICAS ──
    from openpyxl.chart import BarChart
    wst = wb.create_sheet("Estatísticas", 1)

    # Cálculos
    apostas_ganhas   = [a for a in df_res if a["resultado"] == "ganhou"]
    apostas_perdidas = [a for a in df_res if a["resultado"] == "perdeu"]
    capital_em_jogo  = sum(float(a["stake"]) for a in apostas if a["resultado"] == "pendente")
    capital_atual    = BANCA + lucro_total - capital_em_jogo
    maior_odd_acert  = max((float(a["odd"]) for a in apostas_ganhas), default=0)
    maior_stake_val  = max((float(a["stake"]) for a in df_res), default=0)
    stake_total_e    = sum(float(a["stake"]) for a in df_res)
    stake_media_e    = stake_total_e / len(df_res) if df_res else 0
    odd_media_e      = sum(float(a["odd"]) for a in df_res) / len(df_res) if df_res else 0
    maior_lucro_e    = max((lucro_aposta(a) for a in apostas_ganhas), default=0)
    maior_perda_e    = min((lucro_aposta(a) for a in apostas_perdidas), default=0)

    seq_verde_max = seq_verm_max = seq_verde_cur = seq_verm_cur = 0
    for a in df_res:
        if a["resultado"] == "ganhou":
            seq_verde_cur += 1; seq_verm_cur = 0
        elif a["resultado"] == "perdeu":
            seq_verm_cur += 1; seq_verde_cur = 0
        else:
            seq_verde_cur = seq_verm_cur = 0
        seq_verde_max = max(seq_verde_max, seq_verde_cur)
        seq_verm_max  = max(seq_verm_max,  seq_verm_cur)

    # Por Unidade
    por_unidade_e = {}
    for a in df_res:
        u = float(a.get("unidade") or 50)
        if u not in por_unidade_e:
            por_unidade_e[u] = {"ap":0,"g":0,"lucro":0.0,"stake":0.0}
        por_unidade_e[u]["ap"] += 1
        por_unidade_e[u]["lucro"] += lucro_aposta(a)
        por_unidade_e[u]["stake"] += float(a["stake"])
        if a["resultado"] == "ganhou": por_unidade_e[u]["g"] += 1

    # Layout
    wst.merge_cells("A1:C1"); wst["A1"] = "ESTATÍSTICAS"
    est(wst["A1"], bold=True, bg=DARK, size=14)
    wst.row_dimensions[1].height = 34; wst.row_dimensions[2].height = 6

    def stat_row(ws, row, label, valor, fmt=None, cor_val=False):
        cl = ws.cell(row=row, column=1, value=label)
        cl.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cl.fill = PatternFill("solid", start_color="1E3A5F")
        cl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cl.border = brd()
        cv = ws.cell(row=row, column=2, value=valor)
        fc = "000000"
        if cor_val:
            try: fc = GREEN if float(valor) >= 0 else RED
            except: pass
        cv.font = Font(name="Arial", size=10, bold=True, color=fc)
        cv.alignment = Alignment(horizontal="center", vertical="center")
        cv.fill = PatternFill("solid", start_color="EFF6FF")
        cv.border = brd()
        if fmt: cv.number_format = fmt
        ws.row_dimensions[row].height = 20

    metricas = [
        ("Capital Atual (R$)",         round(capital_atual, 2),      "#,##0.00", True),
        ("Stake Investida Total (R$)",  round(stake_total_e, 2),      "#,##0.00", False),
        ("Stake Média (R$)",            round(stake_media_e, 2),      "#,##0.00", False),
        ("Odd Média",                   round(odd_media_e, 3),        "0.000",    False),
        ("Maior Odd Acertada",          round(maior_odd_acert, 3),    "0.000",    False),
        ("Maior Stake (R$)",            round(maior_stake_val, 2),    "#,##0.00", False),
        ("Maior Lucro em uma aposta (R$)", round(maior_lucro_e, 2),  "#,##0.00", True),
        ("Maior Perda em uma aposta (R$)", round(maior_perda_e, 2),  "#,##0.00", True),
        ("Maior Sequência de Greens",   seq_verde_max,                "0",        False),
        ("Maior Sequência de Reds",     seq_verm_max,                 "0",        False),
    ]
    for i,(label,valor,fmt,cor_v) in enumerate(metricas):
        stat_row(wst, 3+i, label, valor, fmt, cor_v)

    # Por Unidade abaixo das métricas
    pu_start = 3 + len(metricas) + 2
    wst.row_dimensions[pu_start-1].height = 8
    wst.merge_cells(f"A{pu_start}:C{pu_start}"); cell_pu = wst[f"A{pu_start}"]
    cell_pu.value = "POR UNIDADE"
    est(cell_pu, bold=True, bg=DARK, size=11)
    wst.row_dimensions[pu_start].height = 22
    for col,h in enumerate(["Unidade","Apostas","Ganhou","Perdeu","Stake","Lucro R$","Lucro Units","ROI","Win Rate"],1):
        cell = wst.cell(row=pu_start+1, column=col, value=h)
        est(cell, bold=True, bg="1E3A5F", size=9); cell.border = brd()
        wst.row_dimensions[pu_start+1].height = 18
    for i,u_val in enumerate(sorted(por_unidade_e.keys())):
        d = por_unidade_e[u_val]
        er = pu_start+2+i; rb = WHITE if i%2==0 else ALT
        roi_u = d["lucro"]/d["stake"] if d["stake"] else 0
        wr_u  = d["g"]/d["ap"] if d["ap"] else 0
        lu_u  = d["lucro"]/u_val if u_val else 0
        for col,val in enumerate([f"R$ {u_val:.0f}/ap",d["ap"],d["g"],d["ap"]-d["g"],round(d["stake"],2),round(d["lucro"],2),round(lu_u,2),roi_u,wr_u],1):
            cell = wst.cell(row=er,column=col,value=val); fc="000000"
            if col in (6,7,8): fc=cor(val)
            cell.font=Font(name="Arial",size=9,color=fc)
            cell.alignment=Alignment(horizontal="center" if col>1 else "left",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if col in (5,6): cell.number_format="#,##0.00"
            if col==7: cell.number_format="+0.00;-0.00;0.00"
            if col in (8,9): cell.number_format="+0.0%;-0.0%;0.0%"
        wst.row_dimensions[er].height=18

    # Larguras
    wst.column_dimensions["A"].width = 32
    wst.column_dimensions["B"].width = 16
    for i in range(3,10): wst.column_dimensions[get_column_letter(i)].width = 12

    # ── Gráficos por faixa de odds (0.5 em 0.5, sem faixas vazias) ──
    odd_max = max((float(a["odd"]) for a in df_res), default=3.5)
    lo = 1.00
    FAIXAS = []
    while lo < odd_max + 0.5:
        hi = round(lo + 0.49, 2)
        FAIXAS.append((round(lo,2), hi))
        lo = round(lo + 0.5, 2)

    def lbl_f(lo,hi): return f"{lo:.2f}-{hi:.2f}"

    dados_faixa = []
    for lo,hi in FAIXAS:
        ap_f = [a for a in df_res if lo <= float(a["odd"]) <= hi]
        if not ap_f: continue
        g_f = sum(1 for a in ap_f if a["resultado"]=="ganhou")
        p_f = sum(1 for a in ap_f if a["resultado"]=="perdeu")
        lucro_f = round(sum(lucro_aposta(a) for a in ap_f),2)
        dados_faixa.append((lbl_f(lo,hi), g_f, p_f, lucro_f))

    # Linha dinâmica: começa após Por Unidade (pu_start + 2 linhas header + nº unidades + 2 de gap)
    n_unidades = len(por_unidade_e)
    data_row = pu_start + 2 + n_unidades + 2  # header PU + col header PU + linhas PU + gap
    hdr_col = 1  # coluna A

    # Título da seção
    wst.merge_cells(f"A{data_row}:D{data_row}")
    cell_fo = wst[f"A{data_row}"]
    cell_fo.value = "POR FAIXA DE ODDS"
    est(cell_fo, bold=True, bg=DARK, size=11)
    wst.row_dimensions[data_row].height = 22
    data_row += 1

    headers_f = ["Faixa de Odds","Ganhou","Perdeu","Lucro R$"]
    for c,h in enumerate(headers_f, hdr_col):
        cell = wst.cell(row=data_row, column=c, value=h)
        est(cell, bold=True, bg="1E3A5F", size=9); cell.border=brd()
        wst.row_dimensions[data_row].height=18
    tbl_hdr_row = data_row
    data_row += 1

    for i,(faixa,g,p,lf) in enumerate(dados_faixa):
        er = data_row+i; rb = WHITE if i%2==0 else ALT
        for c,val in enumerate([faixa,g,p,lf], hdr_col):
            cell=wst.cell(row=er,column=c,value=val)
            fc="000000"
            if c==hdr_col+3: fc=cor(val)
            cell.font=Font(name="Arial",size=9,color=fc)
            cell.alignment=Alignment(horizontal="center" if c>hdr_col else "left",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if c==hdr_col+3: cell.number_format="#,##0.00"
        wst.row_dimensions[er].height=15

    n = len(dados_faixa)
    cats = Reference(wst, min_col=hdr_col,   min_row=tbl_hdr_row+1, max_row=tbl_hdr_row+n)
    d_g  = Reference(wst, min_col=hdr_col+1, min_row=tbl_hdr_row,   max_row=tbl_hdr_row+n)
    d_p  = Reference(wst, min_col=hdr_col+2, min_row=tbl_hdr_row,   max_row=tbl_hdr_row+n)
    d_l  = Reference(wst, min_col=hdr_col+3, min_row=tbl_hdr_row,   max_row=tbl_hdr_row+n)

    # Gráfico 1 — apostas por faixa (coluna F, mesma linha da tabela)
    bar1 = BarChart(); bar1.type="col"; bar1.grouping="clustered"; bar1.overlap=0
    bar1.title="Apostas por Faixa de Odds"; bar1.style=10
    bar1.y_axis.title="Apostas"; bar1.width=24; bar1.height=14
    bar1.add_data(d_g, titles_from_data=True); bar1.add_data(d_p, titles_from_data=True)
    bar1.set_categories(cats)
    bar1.series[0].graphicalProperties.solidFill="16A34A"
    bar1.series[1].graphicalProperties.solidFill="DC2626"
    wst.add_chart(bar1, f"F{tbl_hdr_row}")

    # Gráfico 2 — lucro por faixa (coluna R, mesma linha — lado a lado com o gráfico 1)
    bar2 = BarChart(); bar2.type="col"; bar2.grouping="clustered"
    bar2.title="Lucro R$ por Faixa de Odds"; bar2.style=10
    bar2.y_axis.title="Lucro R$"; bar2.width=24; bar2.height=14
    bar2.add_data(d_l, titles_from_data=True); bar2.set_categories(cats)
    bar2.series[0].graphicalProperties.solidFill="2563EB"
    wst.add_chart(bar2, f"R{tbl_hdr_row}")

    # ── ABA 3: LUCRO POR DIA ──
    from collections import defaultdict
    por_dia=defaultdict(list)
    for a in df_res:
        por_dia[str(a["data"])[:10]].append(a)
    wd=wb.create_sheet("Lucro por Dia")
    wd.merge_cells("A1:D1"); wd["A1"]="LUCRO POR DIA"
    est(wd["A1"],bold=True,bg=DARK,size=14); wd.row_dimensions[1].height=34; wd.row_dimensions[2].height=8
    for c,h in enumerate(["Data","Apostas","Lucro do Dia","Acumulado"],1):
        cell=wd.cell(row=3,column=c,value=h); est(cell,bold=True,bg=DARK,size=10); cell.border=brd()
    wd.row_dimensions[3].height=22
    acum3=0
    for i,(data_k,ap) in enumerate(sorted(por_dia.items())):
        er=4+i; rb=WHITE if i%2==0 else ALT
        lucro_d=sum(lucro_aposta(a) for a in ap)
        acum3+=lucro_d
        try: data_f=__import__("datetime").datetime.strptime(data_k,"%Y-%m-%d").strftime("%d/%m/%Y")
        except: data_f=data_k
        for c,val in enumerate([data_f,len(ap),lucro_d,acum3],1):
            cell=wd.cell(row=er,column=c,value=val); fc="000000"
            if c in (3,4): fc=cor(val)
            cell.font=Font(name="Arial",size=10,color=fc)
            cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if c in (3,4): cell.number_format="#,##0.00"
        wd.row_dimensions[er].height=18
    for i,w in enumerate([14,10,18,18],1): wd.column_dimensions[get_column_letter(i)].width=w

    # ── ABA 3: POR CASA ──
    casas={}
    for a in df_res:
        casa=normalizar_casa(a.get("casa"))
        if casa not in casas: casas[casa]={"ap":0,"g":0,"stake":0.0,"lucro":0.0,"lucro_u":0.0}
        c=casas[casa]; c["ap"]+=1; c["stake"]+=float(a["stake"])
        c["lucro"]+=lucro_aposta(a)
        c["lucro_u"]+=lucro_aposta(a)/float(a.get("unidade") or 50)
        if a["resultado"]=="ganhou": c["g"]+=1
    wc=wb.create_sheet("Por Casa")
    wc.merge_cells("A1:I1"); wc["A1"]="POR CASA"
    est(wc["A1"],bold=True,bg=DARK,size=14); wc.row_dimensions[1].height=34; wc.row_dimensions[2].height=8
    for c,h in enumerate(["Casa","Apostas","Ganhou","Perdeu","Stake","Lucro R$","Lucro Units","ROI","Win Rate"],1):
        cell=wc.cell(row=3,column=c,value=h); est(cell,bold=True,bg=DARK,size=10); cell.border=brd()
    wc.row_dimensions[3].height=22
    for i,(nome,c) in enumerate(sorted(casas.items(),key=lambda x:-x[1]["lucro"])):
        er=4+i; rb=WHITE if i%2==0 else ALT
        roi_c=c["lucro"]/c["stake"] if c["stake"] else 0
        wr_c=c["g"]/c["ap"] if c["ap"] else 0
        for col,val in enumerate([nome,c["ap"],c["g"],c["ap"]-c["g"],c["stake"],c["lucro"],round(c["lucro_u"],2),roi_c,wr_c],1):
            cell=wc.cell(row=er,column=col,value=val); fc="000000"
            if col in (6,7,8): fc=cor(val)
            cell.font=Font(name="Arial",size=10,color=fc)
            cell.alignment=Alignment(horizontal="left" if col==1 else "center",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if col in (5,6): cell.number_format="#,##0.00"
            if col==7: cell.number_format="+0.00;-0.00;0.00"
            if col in (8,9): cell.number_format="+0.0%;-0.0%;0.0%"
        wc.row_dimensions[er].height=18
    for i,w in enumerate([18,10,10,10,14,14,14,10,12],1): wc.column_dimensions[get_column_letter(i)].width=w

    # ── ABA 4: POR ESPORTE ──
    esportes={}
    for a in df_res:
        esp=normalizar_esporte(a.get("esporte")) or "Sem esporte"
        if esp not in esportes: esportes[esp]={"ap":0,"g":0,"stake":0.0,"lucro":0.0,"lucro_u":0.0}
        e=esportes[esp]; e["ap"]+=1; e["stake"]+=float(a["stake"])
        e["lucro"]+=lucro_aposta(a)
        e["lucro_u"]+=lucro_aposta(a)/float(a.get("unidade") or 50)
        if a["resultado"]=="ganhou": e["g"]+=1
    we=wb.create_sheet("Por Esporte")
    we.merge_cells("A1:I1"); we["A1"]="POR ESPORTE"
    est(we["A1"],bold=True,bg=DARK,size=14); we.row_dimensions[1].height=34; we.row_dimensions[2].height=8
    for c,h in enumerate(["Esporte","Apostas","Ganhou","Perdeu","Stake","Lucro R$","Lucro Units","ROI","Win Rate"],1):
        cell=we.cell(row=3,column=c,value=h); est(cell,bold=True,bg=DARK,size=10); cell.border=brd()
    we.row_dimensions[3].height=22
    for i,(nome,e) in enumerate(sorted(esportes.items(),key=lambda x:-x[1]["lucro"])):
        er=4+i; rb=WHITE if i%2==0 else ALT
        roi_e=e["lucro"]/e["stake"] if e["stake"] else 0
        wr_e=e["g"]/e["ap"] if e["ap"] else 0
        for col,val in enumerate([nome,e["ap"],e["g"],e["ap"]-e["g"],e["stake"],e["lucro"],round(e["lucro_u"],2),roi_e,wr_e],1):
            cell=we.cell(row=er,column=col,value=val); fc="000000"
            if col in (6,7,8): fc=cor(val)
            cell.font=Font(name="Arial",size=10,color=fc)
            cell.alignment=Alignment(horizontal="left" if col==1 else "center",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if col in (5,6): cell.number_format="#,##0.00"
            if col==7: cell.number_format="+0.00;-0.00;0.00"
            if col in (8,9): cell.number_format="+0.0%;-0.0%;0.0%"
        we.row_dimensions[er].height=18
    for i,w in enumerate([18,10,10,10,14,14,14,10,12],1): we.column_dimensions[get_column_letter(i)].width=w

    # ── ABA 5: POR SEMANA (domingo a sabado) ──
    import datetime as dt_mod
    def domingo_da_semana(dt):
        """Retorna o domingo que inicia a semana (dom-sab) de uma data."""
        # Garante que seja datetime (date.replace não aceita hour/minute/second)
        if not isinstance(dt, dt_mod.datetime):
            if isinstance(dt, dt_mod.date):
                dt = dt_mod.datetime(dt.year, dt.month, dt.day)
            else:
                dt = dt_mod.datetime.strptime(str(dt)[:10], "%Y-%m-%d")
        dias = (dt.weekday() + 1) % 7
        return (dt - timedelta(days=dias)).replace(hour=0, minute=0, second=0, microsecond=0)
    por_sem=defaultdict(list)
    for a in df_res:
        dom=domingo_da_semana(a["data"])
        por_sem[dom].append(a)
    ws2=wb.create_sheet("Por Semana")
    ws2.merge_cells("A1:I1"); ws2["A1"]="POR SEMANA"
    est(ws2["A1"],bold=True,bg=DARK,size=14); ws2.row_dimensions[1].height=34; ws2.row_dimensions[2].height=8
    for c,h in enumerate(["Semana","Periodo","Apostas","Ganhou","Perdeu","Stake","Lucro R$","Lucro Units","ROI"],1):
        cell=ws2.cell(row=3,column=c,value=h); est(cell,bold=True,bg=DARK,size=10); cell.border=brd()
    ws2.row_dimensions[3].height=22
    for i,(dom,ap) in enumerate(sorted(por_sem.items())):
        er=4+i; rb=WHITE if i%2==0 else ALT
        lucro_s=sum(lucro_aposta(a) for a in ap)
        stake_s=sum(float(a["stake"]) for a in ap)
        g_s=sum(1 for a in ap if a["resultado"]=="ganhou")
        roi_s=lucro_s/stake_s if stake_s else 0
        lucro_u_s=sum(lucro_aposta(a)/float(a.get("unidade") or 50) for a in ap)
        sab=dom+timedelta(days=6)
        num_semana=dom.isocalendar()[1]
        periodo=f"{dom.strftime('%d/%m')} - {sab.strftime('%d/%m/%Y')}"
        for col,val in enumerate([f"Semana {num_semana}",periodo,len(ap),g_s,len(ap)-g_s,stake_s,lucro_s,round(lucro_u_s,2),roi_s],1):
            cell=ws2.cell(row=er,column=col,value=val); fc="000000"
            if col in (7,8,9): fc=cor(val)
            cell.font=Font(name="Arial",size=10,color=fc)
            cell.alignment=Alignment(horizontal="left" if col==2 else "center",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if col in (6,7): cell.number_format="#,##0.00"
            if col==8: cell.number_format="+0.00;-0.00;0.00"
            if col==9: cell.number_format="+0.0%;-0.0%;0.0%"
        ws2.row_dimensions[er].height=18
    for i,w in enumerate([12,22,10,10,10,14,14,14,10],1): ws2.column_dimensions[get_column_letter(i)].width=w

    # ── ABA 6: POR MÊS ──
    from collections import defaultdict as _dd2
    NOMES_MES_XL = ["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    por_mes_xl = _dd2(list)
    for a in df_res:
        d = a["data"] if hasattr(a["data"],"strftime") else __import__("datetime").datetime.strptime(str(a["data"])[:10],"%Y-%m-%d")
        por_mes_xl[d.strftime("%Y-%m")].append(a)

    wm = wb.create_sheet("Por Mes")
    wm.merge_cells("A1:I1"); wm["A1"] = "POR MÊS"
    est(wm["A1"],bold=True,bg=DARK,size=14); wm.row_dimensions[1].height=34; wm.row_dimensions[2].height=8
    for c,h in enumerate(["Mês","Apostas","Ganhou","Perdeu","Stake","Lucro R$","Lucro Units","ROI","Win Rate"],1):
        cell=wm.cell(row=3,column=c,value=h); est(cell,bold=True,bg=DARK,size=10); cell.border=brd()
    wm.row_dimensions[3].height=22

    acum_m = 0
    for i,chave in enumerate(sorted(por_mes_xl.keys())):
        ap_m = por_mes_xl[chave]
        ano_m,mes_m = int(chave[:4]),int(chave[5:])
        lucro_m  = sum(lucro_aposta(a) for a in ap_m)
        stake_m  = sum(float(a["stake"]) for a in ap_m)
        g_m      = sum(1 for a in ap_m if a["resultado"]=="ganhou")
        roi_m    = lucro_m/stake_m if stake_m else 0
        wr_m     = g_m/len(ap_m) if ap_m else 0
        # Lucro em unidades: cada aposta usa a unidade registrada nela
        lucro_u_m = sum(lucro_aposta(a)/float(a.get("unidade") or 50) for a in ap_m)
        acum_m  += lucro_m
        er=4+i; rb=WHITE if i%2==0 else ALT
        for col,val in enumerate([f"{NOMES_MES_XL[mes_m]}/{ano_m}",len(ap_m),g_m,len(ap_m)-g_m,stake_m,lucro_m,round(lucro_u_m,2),roi_m,wr_m],1):
            cell=wm.cell(row=er,column=col,value=val); fc="000000"
            if col in (6,7,8): fc=cor(val)
            cell.font=Font(name="Arial",size=10,color=fc)
            cell.alignment=Alignment(horizontal="left" if col==1 else "center",vertical="center")
            cell.fill=PatternFill("solid",start_color=rb); cell.border=brd()
            if col in (5,6): cell.number_format="#,##0.00"
            if col==7: cell.number_format="+0.00;-0.00;0.00"
            if col in (8,9): cell.number_format="+0.0%;-0.0%;0.0%"
        wm.row_dimensions[er].height=18

    # Linha de total
    tr_m = 4+len(por_mes_xl)
    for col in range(1,10):
        cell=wm.cell(row=tr_m,column=col)
        if col==1: cell.value="TOTAL"
        elif col==2: cell.value=len(df_res)
        elif col==3: cell.value=sum(1 for a in df_res if a["resultado"]=="ganhou")
        elif col==4: cell.value=sum(1 for a in df_res if a["resultado"]=="perdeu")
        elif col==5: cell.value=round(sum(float(a["stake"]) for a in df_res),2); cell.number_format="#,##0.00"
        elif col==6: cell.value=round(sum(lucro_aposta(a) for a in df_res),2); cell.number_format="#,##0.00"; fc=cor(cell.value); cell.font=Font(name="Arial",bold=True,size=10,color=fc)
        elif col==7:
            total_u=sum(lucro_aposta(a)/float(a.get("unidade") or 50) for a in df_res)
            cell.value=round(total_u,2); cell.number_format="+0.00;-0.00;0.00"; fc=cor(cell.value); cell.font=Font(name="Arial",bold=True,size=10,color=fc)
        if col not in (6,7): cell.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
        est(cell,bold=True,bg=DARK,size=10); cell.border=brd()
    wm.row_dimensions[tr_m].height=22
    for i,w in enumerate([12,10,10,10,14,14,14,10,12],1): wm.column_dimensions[get_column_letter(i)].width=w

    # ── ABA 7: GRAFICO ──
    wg=wb.create_sheet("Evolucao da Banca")
    wg["A1"]="Aposta #"; wg["B1"]="Banca Acumulada"
    for i,(val) in enumerate(banca_acum,2):
        wg.cell(row=i,column=1,value=i-1); wg.cell(row=i,column=2,value=round(val,2))
    if len(banca_acum)>=2:
        chart=LineChart(); chart.title="Evolucao da Banca"; chart.style=10
        chart.y_axis.title="R$"; chart.x_axis.title="Aposta #"
        chart.y_axis.numFmt="#,##0.00"; chart.width=26; chart.height=14
        dr=Reference(wg,min_col=2,min_row=1,max_row=len(banca_acum)+1)
        cr=Reference(wg,min_col=1,min_row=2,max_row=len(banca_acum)+1)
        chart.add_data(dr,titles_from_data=True); chart.set_categories(cr)
        chart.series[0].graphicalProperties.line.solidFill="2563EB"
        chart.series[0].graphicalProperties.line.width=22000
        wg.add_chart(chart,"D2")

    # Salvar na memória e enviar
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import datetime as dt2
    nome_arq=f"dashboard_{dt2.now().strftime('%d%m%Y_%H%M')}.xlsx"
    await update.message.reply_document(
        document=InputFile(buf, filename=nome_arq),
        caption=f"Dashboard gerado! {total} apostas | R$ {lucro_total:.2f} lucro | ROI {roi:+.1%}"
    )

# ── EXPORTAR CSV ──────────────────────────────────────────────────────────────
async def exportar_csv(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    apostas   = carregar()
    total     = len(apostas)
    pendentes = sum(1 for a in apostas if a["resultado"] == "pendente")
    import io
    output = io.StringIO()
    import csv as csv_mod
    writer = csv_mod.writer(output, quoting=csv_mod.QUOTE_ALL)
    writer.writerow(["id","data","horario","descricao","odd","stake","resultado","casa","esporte"])
    for a in apostas:
        data = a["data"].strftime("%Y-%m-%d") if hasattr(a["data"],"strftime") else str(a["data"])
        writer.writerow([a["id"],data,a.get("horario",""),a["descricao"],a["odd"],a["stake"],a["resultado"],a.get("casa",""),a.get("esporte","")])
    csv_bytes = output.getvalue().encode("utf-8")
    caption   = f"\U0001f4ca apostas.csv\n{total} apostas | {pendentes} pendentes"
    await update.message.reply_document(document=InputFile(csv_bytes, filename="apostas.csv"), caption=caption)

# ── SERVIDOR KEEP-ALIVE ───────────────────────────────────────────────────────
class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers(); self.wfile.write(b"OK")
    def do_HEAD(self):
        self.send_response(200); self.end_headers()
    def log_message(self, *args): pass

def iniciar_servidor():
    port = int(os.environ.get("PORT", 8080))
    HTTPServer(("0.0.0.0", port), PingHandler).serve_forever()

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    inicializar_db()
    app = Application.builder().token(TOKEN).build()

    BOTOES_MENU = ["📝 Nova aposta","⏳ Ver pendentes","📈 Resultados","✏️ Editar aposta","📊 Gerar Resultados","⚙️ Mudar Unidade"]
    fallbacks_padrao = [
        CommandHandler("cancelar", cancelar),
        MessageHandler(filters.Regex(f"^{CANCELAR_BTN}$"), cancelar),
        MessageHandler(filters.Regex("^(" + "|".join(re.escape(b) for b in BOTOES_MENU) + ")$"), cancelar),
    ]

    conv_nova = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📝 Nova aposta$"), nova_aposta_inicio)],
        states={
            DATA:      [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_data)],
            HORARIO:   [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_horario)],
            DESCRICAO: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_descricao)],
            ODD:       [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_odd)],
            STAKE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_stake)],
            ESPORTE:   [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_esporte)],
            CASA:      [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_casa)],
        },
        fallbacks=fallbacks_padrao,
    )

    conv_editar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^✏️ Editar aposta$"), editar_inicio)],
        states={
            EDITAR_ID:    [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_receber_id)],
            EDITAR_CAMPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_receber_campo)],
            EDITAR_VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_receber_valor)],
            EDITAR_CASA:  [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_receber_casa)],
        },
        fallbacks=fallbacks_padrao,
    )

    conv_unidade = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^⚙️ Mudar Unidade$"), mudar_unidade_inicio)],
        states={
            MUDAR_UNIDADE_VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, mudar_unidade_receber)],
        },
        fallbacks=fallbacks_padrao,
    )

    conv_resultados = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📈 Resultados$"), resultados)],
        states={
            RESULTADOS_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, resultados_resposta)],
        },
        fallbacks=fallbacks_padrao,
    )

    conv_gerar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📊 Gerar Resultados$"), gerar_menu)],
        states={
            GERAR_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, gerar_resposta)],
        },
        fallbacks=fallbacks_padrao,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("exportar", exportar_csv))
    app.add_handler(CommandHandler("migrar_casas", cmd_migrar_casas))
    app.add_handler(conv_nova)
    app.add_handler(conv_editar)
    app.add_handler(conv_unidade)
    app.add_handler(conv_resultados)
    app.add_handler(conv_gerar)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_botao))

    threading.Thread(target=iniciar_servidor, daemon=True).start()
    print("🤖 Bot rodando! Abra o Telegram e mande /start")
    app.run_polling()

if __name__ == "__main__":
    main()
