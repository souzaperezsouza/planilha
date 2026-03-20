import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import sys
import os

CSV_FILE = "apostas.csv"
OUTPUT_FILE = "dashboard_apostas.xlsx"
BANCA_INICIAL = 5000

if not os.path.exists(CSV_FILE):
    print(f"Arquivo '{CSV_FILE}' não encontrado.")
    sys.exit(1)

df = pd.read_csv(CSV_FILE, parse_dates=["data"])
df.columns = df.columns.str.strip().str.lower()
df["resultado"] = df["resultado"].str.strip().str.lower()
if "horario" not in df.columns:
    df["horario"] = ""
df["horario"] = df["horario"].fillna("")
if "casa"    not in df.columns: df["casa"]    = ""
if "esporte" not in df.columns: df["esporte"] = ""
df["casa"]    = df["casa"].fillna("").str.strip()
df["esporte"] = df["esporte"].fillna("").str.strip()
df = df.sort_values("data").reset_index(drop=True)

df_res = df[df["resultado"].isin(["ganhou", "perdeu"])].copy()
df_res["lucro"] = df_res.apply(
    lambda r: r["stake"] * (r["odd"] - 1) if r["resultado"] == "ganhou" else -r["stake"], axis=1
)
df_res["banca_acum"] = df_res["lucro"].cumsum()
df_res["progressao"] = df_res["banca_acum"] / BANCA_INICIAL

total      = len(df)
resolvidas = len(df_res)
pendentes  = total - resolvidas
vitorias   = (df_res["resultado"] == "ganhou").sum()
win_rate   = vitorias / resolvidas if resolvidas else 0
total_stake = df_res["stake"].sum()
lucro_total = df_res["lucro"].sum()
roi         = lucro_total / total_stake if total_stake else 0
progressao  = lucro_total / BANCA_INICIAL

lucro_dia = df_res.groupby("data")["lucro"].sum().reset_index()
lucro_dia.columns = ["data", "lucro_dia"]

# Por casa
lucro_casa = df_res[df_res["casa"] != ""].groupby("casa").agg(
    apostas=("lucro", "count"),
    ganhou=("resultado", lambda x: (x == "ganhou").sum()),
    perdeu=("resultado", lambda x: (x == "perdeu").sum()),
    stake=("stake", "sum"),
    lucro=("lucro", "sum"),
).reset_index()
lucro_casa["roi"] = lucro_casa["lucro"] / lucro_casa["stake"]
lucro_casa["win_rate"] = lucro_casa["ganhou"] / lucro_casa["apostas"]
lucro_casa = lucro_casa.sort_values("lucro", ascending=False).reset_index(drop=True)

# ── Estilos ──────────────────────────────────────────────────────────────────
DARK_BG  = "1E293B"
GREEN_BG = "16A34A"
RED_BG   = "DC2626"
AMBER_BG = "D97706"
ALT_BG   = "EFF6FF"
WHITE    = "FFFFFF"
BORDER_C = "CBD5E1"

def estilo(cell, bold=False, fc=WHITE, bg=None, size=11, align="center"):
    cell.font = Font(name="Arial", bold=bold, color=fc, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def borda():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def cor_valor(val):
    try:
        return GREEN_BG if float(val) >= 0 else RED_BG
    except:
        return "000000"

wb = openpyxl.Workbook()

# ── ABA 1: DASHBOARD ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Dashboard"

ws.merge_cells("A1:J1")
ws["A1"] = "📊  DASHBOARD DE APOSTAS"
estilo(ws["A1"], bold=True, bg=DARK_BG, size=16)
ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 8

cards = [
    ("Total Apostas", str(total),                   DARK_BG),
    ("Resolvidas",    str(resolvidas),               DARK_BG),
    ("Pendentes",     str(pendentes),                AMBER_BG if pendentes else DARK_BG),
    ("Win Rate",      f"{win_rate:.1%}",             GREEN_BG if win_rate >= 0.5 else RED_BG),
    ("Lucro Total",   f"R$ {lucro_total:.2f}",       GREEN_BG if lucro_total >= 0 else RED_BG),
    ("ROI",           f"{roi:.1%}",                  GREEN_BG if roi >= 0 else RED_BG),
    ("Progressão",    f"{progressao:+.2%}",          GREEN_BG if progressao >= 0 else RED_BG),
]

for i, (label, val, bg) in enumerate(cards, 1):
    lbl = ws.cell(row=3, column=i, value=label)
    estilo(lbl, bg=bg, size=9, fc="DBEAFE")
    ws.row_dimensions[3].height = 20
    v = ws.cell(row=4, column=i, value=val)
    estilo(v, bold=True, bg=bg, size=14)
    ws.row_dimensions[4].height = 34
    d = ws.cell(row=5, column=i)
    estilo(d, bg="0F172A")
    ws.row_dimensions[5].height = 4

ws.row_dimensions[6].height = 10

HDR_ROW  = 7
DATA_ROW = 8
headers = ["#", "Data", "Hora", "Descrição", "Odd", "Stake (R$)", "Esporte", "Casa", "Resultado", "Lucro (R$)", "Banca Acum.", "Progressão"]
ws.row_dimensions[HDR_ROW].height = 24
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()

for i, row in df.iterrows():
    er = DATA_ROW + i
    ws.row_dimensions[er].height = 18
    rb = WHITE if i % 2 == 0 else ALT_BG

    match_res = df_res[df_res["id"] == row["id"]]
    lucro_val  = match_res.iloc[0]["lucro"]       if not match_res.empty else ""
    banca_val  = match_res.iloc[0]["banca_acum"]  if not match_res.empty else ""
    prog_val   = match_res.iloc[0]["progressao"]  if not match_res.empty else ""

    res_display = {"ganhou": "✅ Ganhou", "perdeu": "❌ Perdeu", "void": "↩️ Void", "pendente": "⏳ Pendente"}.get(row["resultado"], row["resultado"])

    horario_val = str(row.get("horario", ""))[:5] if str(row.get("horario", "")).strip() else ""
    vals = [row["id"], row["data"].strftime("%d/%m/%Y"), horario_val, row["descricao"],
            row["odd"], row["stake"], row.get("esporte",""), row["casa"], res_display, lucro_val, banca_val, prog_val]

    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=er, column=c, value=val)
        fc = "000000"
        if c in (10, 11) and val != "":
            fc = cor_valor(val)
        if c == 12 and val != "":
            fc = cor_valor(val)
        cell.font = Font(name="Arial", size=10, color=fc)
        cell.alignment = Alignment(horizontal="left" if c in (4,) else "center", vertical="center")
        cell.fill = PatternFill("solid", start_color=rb)
        cell.border = borda()
        if c in (5, 6, 10, 11) and val != "":
            cell.number_format = '#,##0.00'
        if c == 12 and val != "":
            cell.number_format = '+0.00%;-0.00%;0.00%'

tr = DATA_ROW + len(df)
ws.row_dimensions[tr].height = 22
for c in range(1, 13):
    cell = ws.cell(row=tr, column=c)
    if c == 4:
        cell.value = "TOTAL"
    elif c == 6:
        cell.value = f"=SUM(F{DATA_ROW}:F{tr-1})"
        cell.number_format = '#,##0.00'
    elif c == 10:
        cell.value = f"=SUM(J{DATA_ROW}:J{tr-1})"
        cell.number_format = '#,##0.00'
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()

widths = [5, 12, 7, 34, 7, 11, 14, 14, 14, 12, 14, 13]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── ABA 2: LUCRO POR DIA ─────────────────────────────────────────────────────
wd = wb.create_sheet("Lucro por Dia")
wd.merge_cells("A1:D1")
wd["A1"] = "📅  LUCRO POR DIA"
estilo(wd["A1"], bold=True, bg=DARK_BG, size=14)
wd.row_dimensions[1].height = 36
wd.row_dimensions[2].height = 8

for c, h in enumerate(["Data", "Apostas", "Lucro do Dia (R$)", "Acumulado (R$)"], 1):
    cell = wd.cell(row=3, column=c, value=h)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
wd.row_dimensions[3].height = 22

acum = 0
for i, row in lucro_dia.iterrows():
    er = 4 + i
    acum += row["lucro_dia"]
    rb = WHITE if i % 2 == 0 else ALT_BG
    n_ap = len(df_res[df_res["data"] == row["data"]])
    for c, val in enumerate([row["data"].strftime("%d/%m/%Y"), n_ap, row["lucro_dia"], acum], 1):
        cell = wd.cell(row=er, column=c, value=val)
        fc = "000000"
        if c in (3, 4):
            fc = cor_valor(val)
        cell.font = Font(name="Arial", size=10, color=fc)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color=rb)
        cell.border = borda()
        if c in (3, 4):
            cell.number_format = '#,##0.00'
    wd.row_dimensions[er].height = 18

for i, w in enumerate([14, 10, 20, 18], 1):
    wd.column_dimensions[get_column_letter(i)].width = w

# ── ABA 3: POR CASA ──────────────────────────────────────────────────────────
wc = wb.create_sheet("Por Casa")
wc.merge_cells("A1:G1")
wc["A1"] = "🏦  DESEMPENHO POR CASA DE APOSTA"
estilo(wc["A1"], bold=True, bg=DARK_BG, size=14)
wc.row_dimensions[1].height = 36
wc.row_dimensions[2].height = 8

hdrs_casa = ["Casa", "Apostas", "Ganhou", "Perdeu", "Stake (R$)", "Lucro (R$)", "ROI", "Win Rate"]
for c, h in enumerate(hdrs_casa, 1):
    cell = wc.cell(row=3, column=c, value=h)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
wc.row_dimensions[3].height = 22

for i, row in lucro_casa.iterrows():
    er = 4 + i
    rb = WHITE if i % 2 == 0 else ALT_BG
    vals = [row["casa"], int(row["apostas"]), int(row["ganhou"]), int(row["perdeu"]),
            row["stake"], row["lucro"], row["roi"], row["win_rate"]]
    for c, val in enumerate(vals, 1):
        cell = wc.cell(row=er, column=c, value=val)
        fc = "000000"
        if c == 6:
            fc = cor_valor(val)
        if c == 7:
            fc = cor_valor(val)
        if c == 8:
            fc = GREEN_BG if float(val) >= 0.5 else RED_BG
        cell.font = Font(name="Arial", size=10, color=fc)
        cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
        cell.fill = PatternFill("solid", start_color=rb)
        cell.border = borda()
        if c in (5, 6):
            cell.number_format = '#,##0.00'
        if c in (7, 8):
            cell.number_format = '+0.0%;-0.0%;0.0%'
    wc.row_dimensions[er].height = 18

# Total
tr_c = 4 + len(lucro_casa)
wc.row_dimensions[tr_c].height = 22
total_stake_c = lucro_casa["stake"].sum()
total_lucro_c = lucro_casa["lucro"].sum()
total_ap_c    = lucro_casa["apostas"].sum()
total_g_c     = lucro_casa["ganhou"].sum()
total_p_c     = lucro_casa["perdeu"].sum()
roi_total_c   = total_lucro_c / total_stake_c if total_stake_c else 0
wr_total_c    = total_g_c / total_ap_c if total_ap_c else 0
totais_c = ["TOTAL", int(total_ap_c), int(total_g_c), int(total_p_c),
            total_stake_c, total_lucro_c, roi_total_c, wr_total_c]
for c, val in enumerate(totais_c, 1):
    cell = wc.cell(row=tr_c, column=c, value=val)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
    if c in (5, 6):
        cell.number_format = '#,##0.00'
    if c in (7, 8):
        cell.number_format = '+0.0%;-0.0%;0.0%'

for i, w in enumerate([18, 10, 10, 10, 14, 14, 10, 12], 1):
    wc.column_dimensions[get_column_letter(i)].width = w

# ── ABA 4: EVOLUÇÃO DA BANCA (dados + gráfico com data labels) ───────────────
wg = wb.create_sheet("Evolução da Banca")

for c, h in enumerate(["Aposta #", "Banca Acum. (R$)", "Progressão (%)"], 1):
    cell = wg.cell(row=1, column=c, value=h)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
wg.row_dimensions[1].height = 22

for i, (_, row) in enumerate(df_res.iterrows(), 2):
    banca = round(float(row["banca_acum"]), 2)
    prog  = round(float(row["progressao"]) * 100, 2)
    rb = WHITE if i % 2 == 0 else ALT_BG
    for c, val in enumerate([i - 1, banca, prog], 1):
        cell = wg.cell(row=i, column=c, value=val)
        fc = "000000"
        if c in (2, 3):
            fc = cor_valor(val)
        cell.font = Font(name="Arial", size=10, color=fc)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color=rb)
        cell.border = borda()
        if c == 2:
            cell.number_format = '#,##0.00'
        if c == 3:
            cell.number_format = '+0.00;-0.00;0.00'
    wg.row_dimensions[i].height = 18

for i, w in enumerate([12, 20, 16], 1):
    wg.column_dimensions[get_column_letter(i)].width = w

if len(df_res) >= 2:
    n = len(df_res) + 1

    def make_chart(width, height, show_labels=False):
        c = LineChart()
        c.title = "Evolução da Banca"
        c.style = 10
        c.y_axis.title = "Lucro Acumulado (R$)"
        c.x_axis.title = "Nº da Aposta"
        c.y_axis.numFmt = '#,##0.00'
        c.y_axis.majorGridlines = None
        c.width  = width
        c.height = height
        dr = Reference(wg, min_col=2, min_row=1, max_row=n)
        cr = Reference(wg, min_col=1, min_row=2, max_row=n)
        c.add_data(dr, titles_from_data=True)
        c.set_categories(cr)
        s = c.series[0]
        s.graphicalProperties.line.solidFill = "2563EB"
        s.graphicalProperties.line.width = 22000
        if show_labels:
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True
            s.dLbls.showLegendKey = False
            s.dLbls.showCatName = False
            s.dLbls.showSerName = False
        return c

    wg.add_chart(make_chart(28, 16, show_labels=True), "E2")
    ws.add_chart(make_chart(24, 14, show_labels=False), f"A{tr + 2}")


# ── ABA 5: POR SEMANA (Dom->Sab, numeracao igual ao Betanalytix) ─────────────
from datetime import timedelta

def semana_betanalytix(dt):
    domingo = dt - timedelta(days=(dt.weekday() + 1) % 7)
    return domingo.isocalendar()[0], domingo.isocalendar()[1]

df_res2 = df_res.copy()
df_res2["semana_ano"] = df_res2["data"].apply(lambda d: semana_betanalytix(d)[0])
df_res2["semana_num"] = df_res2["data"].apply(lambda d: semana_betanalytix(d)[1])
df_res2["semana_key"] = df_res2["data"].apply(
    lambda d: semana_betanalytix(d)[0] * 100 + semana_betanalytix(d)[1]
)

lucro_semana = df_res2.groupby(["semana_key","semana_ano","semana_num"]).agg(
    apostas=("lucro","count"),
    ganhou=("resultado", lambda x: (x=="ganhou").sum()),
    stake=("stake","sum"),
    lucro=("lucro","sum"),
).reset_index().sort_values("semana_key")

def domingo_da_semana(ano, semana):
    from datetime import datetime as _dt
    jan1 = _dt(ano, 1, 1)
    iso1 = jan1.isocalendar()
    primeira_seg = jan1 - timedelta(days=iso1[2]-1)
    seg_da_sem   = primeira_seg + timedelta(weeks=semana-1)
    return seg_da_sem - timedelta(days=1)

ws2 = wb.create_sheet("Por Semana")
ws2.merge_cells("A1:H1")
ws2["A1"] = "\U0001f4c5  DESEMPENHO POR SEMANA"
estilo(ws2["A1"], bold=True, bg=DARK_BG, size=14)
ws2.row_dimensions[1].height = 36
ws2.row_dimensions[2].height = 8

hdrs_sem = ["Semana","Periodo","Apostas","Ganhou","Perdeu","Stake (R$)","Lucro (R$)","ROI"]
for c, h in enumerate(hdrs_sem, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
ws2.row_dimensions[3].height = 22

for i, row in lucro_semana.iterrows():
    er   = 4 + i
    rb   = WHITE if i % 2 == 0 else ALT_BG
    ano  = int(row["semana_ano"])
    num  = int(row["semana_num"])
    dom  = domingo_da_semana(ano, num)
    sab  = dom + timedelta(days=6)
    periodo  = f"{dom.strftime('%d/%m')} - {sab.strftime('%d/%m/%Y')}"
    roi_sem  = row["lucro"] / row["stake"] if row["stake"] else 0
    perdeu   = int(row["apostas"]) - int(row["ganhou"])
    vals = [f"Semana {num}", periodo, int(row["apostas"]), int(row["ganhou"]),
            perdeu, row["stake"], row["lucro"], roi_sem]
    for c, val in enumerate(vals, 1):
        cell = ws2.cell(row=er, column=c, value=val)
        fc = "000000"
        if c == 7: fc = cor_valor(val)
        if c == 8: fc = cor_valor(val)
        cell.font = Font(name="Arial", size=10, color=fc)
        cell.alignment = Alignment(horizontal="left" if c == 2 else "center", vertical="center")
        cell.fill = PatternFill("solid", start_color=rb)
        cell.border = borda()
        if c in (6, 7): cell.number_format = "#,##0.00"
        if c == 8:      cell.number_format = "+0.0%;-0.0%;0.0%"
    ws2.row_dimensions[er].height = 18

tr_s    = 4 + len(lucro_semana)
ws2.row_dimensions[tr_s].height = 22
tot_ap  = lucro_semana["apostas"].sum()
tot_g   = lucro_semana["ganhou"].sum()
tot_st  = lucro_semana["stake"].sum()
tot_lu  = lucro_semana["lucro"].sum()
roi_tot = tot_lu / tot_st if tot_st else 0
for c, val in enumerate(["TOTAL","",int(tot_ap),int(tot_g),int(tot_ap-tot_g),tot_st,tot_lu,roi_tot], 1):
    cell = ws2.cell(row=tr_s, column=c, value=val)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
    if c in (6, 7): cell.number_format = "#,##0.00"
    if c == 8:      cell.number_format = "+0.0%;-0.0%;0.0%"

for i, w in enumerate([12, 22, 10, 10, 10, 14, 14, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ── ABA: POR ESPORTE ─────────────────────────────────────────────────────────
lucro_esporte = df_res[df_res["esporte"] != ""].groupby("esporte").agg(
    apostas=("lucro","count"),
    ganhou=("resultado", lambda x: (x=="ganhou").sum()),
    perdeu=("resultado", lambda x: (x=="perdeu").sum()),
    stake=("stake","sum"),
    lucro=("lucro","sum"),
).reset_index()
lucro_esporte["roi"]      = lucro_esporte["lucro"] / lucro_esporte["stake"]
lucro_esporte["win_rate"] = lucro_esporte["ganhou"] / lucro_esporte["apostas"]
lucro_esporte = lucro_esporte.sort_values("lucro", ascending=False).reset_index(drop=True)

we = wb.create_sheet("Por Esporte")
we.merge_cells("A1:H1")
we["A1"] = "🏅  DESEMPENHO POR ESPORTE"
estilo(we["A1"], bold=True, bg=DARK_BG, size=14)
we.row_dimensions[1].height = 36
we.row_dimensions[2].height = 8

for c, h in enumerate(["Esporte","Apostas","Ganhou","Perdeu","Stake (R$)","Lucro (R$)","ROI","Win Rate"], 1):
    cell = we.cell(row=3, column=c, value=h)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
we.row_dimensions[3].height = 22

for i, row in lucro_esporte.iterrows():
    er = 4 + i
    rb = WHITE if i % 2 == 0 else ALT_BG
    vals = [row["esporte"], int(row["apostas"]), int(row["ganhou"]), int(row["perdeu"]),
            row["stake"], row["lucro"], row["roi"], row["win_rate"]]
    for c, val in enumerate(vals, 1):
        cell = we.cell(row=er, column=c, value=val)
        fc = "000000"
        if c == 6: fc = cor_valor(val)
        if c == 7: fc = cor_valor(val)
        if c == 8: fc = GREEN_BG if float(val) >= 0.5 else RED_BG
        cell.font = Font(name="Arial", size=10, color=fc)
        cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
        cell.fill = PatternFill("solid", start_color=rb)
        cell.border = borda()
        if c in (5, 6): cell.number_format = "#,##0.00"
        if c in (7, 8): cell.number_format = "+0.0%;-0.0%;0.0%"
    we.row_dimensions[er].height = 18

tr_e = 4 + len(lucro_esporte)
we.row_dimensions[tr_e].height = 22
tot_ap = lucro_esporte["apostas"].sum()
tot_g  = lucro_esporte["ganhou"].sum()
tot_p  = lucro_esporte["perdeu"].sum()
tot_st = lucro_esporte["stake"].sum()
tot_lu = lucro_esporte["lucro"].sum()
roi_t  = tot_lu / tot_st if tot_st else 0
wr_t   = tot_g / tot_ap if tot_ap else 0
for c, val in enumerate(["TOTAL",int(tot_ap),int(tot_g),int(tot_p),tot_st,tot_lu,roi_t,wr_t], 1):
    cell = we.cell(row=tr_e, column=c, value=val)
    estilo(cell, bold=True, bg=DARK_BG, size=10)
    cell.border = borda()
    if c in (5, 6): cell.number_format = "#,##0.00"
    if c in (7, 8): cell.number_format = "+0.0%;-0.0%;0.0%"

for i, w in enumerate([18, 10, 10, 10, 14, 14, 10, 12], 1):
    we.column_dimensions[get_column_letter(i)].width = w

wb.save(OUTPUT_FILE)
print(f"\n  ✅ Dashboard gerado: {OUTPUT_FILE}\n")
